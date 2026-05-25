#!/usr/bin/env python3
"""
电商图片批处理工具 - PySide6 独立 EXE 版
双击运行，无需额外文件。
依赖：pip install PySide6 Pillow openpyxl
打包：pyinstaller --onefile --windowed --collect-all PIL --collect-all openpyxl --collect-all PySide6 barcode_image_mover_exe.py
"""

import os
import sys
import re
import shutil
import zipfile
import datetime
import threading
import base64
import tempfile
import io
import csv
import uuid
import atexit
from concurrent.futures import ThreadPoolExecutor, wait, FIRST_COMPLETED
from typing import NamedTuple, Optional
import winsound
from PIL import Image, ImageOps
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                               QHBoxLayout, QLabel, QLineEdit,
                               QPushButton, QComboBox, QCheckBox, QTextEdit,
                               QFileDialog, QMessageBox, QGridLayout,
                               QProgressBar, QFrame, QGraphicsDropShadowEffect,
                               QSizePolicy, QListView, QTabWidget)
from PySide6.QtCore import Qt, QThread, Signal, QObject, QTimer, QSize, QRectF
from PySide6.QtGui import (QTextCursor, QTextCharFormat, QColor, QFont, QIcon, QPixmap,
                            QPalette, QPainter, QLinearGradient, QPainterPath)


# ============================================================
# ============================================================
# 核心模块（内嵌，无外部依赖）
# ============================================================

from app.constants import (
    IMAGE_EXTS, TARGET_SIZE, ZIP_SPLIT_BYTES, COPY_WORKERS, IMAGE_WORKERS,
    MAIN_IMAGE_MAX_BYTES, DETAIL_IMAGE_MAX_BYTES,
    MAIN_JPEG_QUALITY, MAIN_JPEG_COMPRESS_START_QUALITY, MAIN_JPEG_MIN_QUALITY,
    DETAIL_JPEG_OPTIMIZE_QUALITY, DETAIL_JPEG_QUALITIES, DETAIL_SCALE_FACTORS,
    EXCEL_EMPTY_ROW_BREAK_THRESHOLD, MANUAL_REVIEW_DIR_NAME,
    LOG_FLUSH_INTERVAL, PROGRESS_REPORT_FRACTION, EXECUTOR_POLL_SECONDS,
    LOG_MAX_BLOCK_COUNT, DEBOUNCE_MS, SHIMMER_INTERVAL_MS,
    SHIMMER_STEP, SHIMMER_BAND_WIDTH, RESAMPLE_LANCZOS,
    MODE_NAMES, RESIZE_MODE_MAP, MODE_HELP_TEXT, UI_HELP_TEXT,
)
from app.services.logger import LogWriter
from app.services.icons import _darken, _write_temp_svg, _TEMP_ICON_PATHS, _cleanup_temp_icons

_UNIQUE_PATH_LOCK = threading.Lock()
_RESERVED_OUTPUT_PATHS = set()


class MainImageResult(NamedTuple):
    filename: str
    ok: bool
    error: Optional[str]
    manual_copy: Optional[str] = None


class DetailImageResult(NamedTuple):
    filename: str
    ok: bool
    error: Optional[str]
    converted: int
    compressed: int
    info: Optional[str]
    manual_copy: Optional[str] = None


def clean_detail_suffix(filename):
    name, ext = os.path.splitext(filename)
    new = name.replace('_详情图', '')
    return new + ext if new != name else filename


def is_image_file(filename):
    return os.path.splitext(filename)[1].lower() in IMAGE_EXTS


def iter_files(folder):
    with os.scandir(folder) as entries:
        for entry in entries:
            if entry.is_file():
                yield entry.name


def iter_image_files(folder):
    return (name for name in iter_files(folder) if is_image_file(name))


def split_source_images(source_dir):
    """一次扫描源目录，分离详情图/主图列表，减少大目录重复扫描。"""
    detail_files = []
    main_files = []
    for fn in iter_image_files(source_dir):
        if '_详情图' in fn:
            detail_files.append(fn)
        else:
            main_files.append(fn)
    return detail_files, main_files


def dir_has_files(folder):
    if not os.path.isdir(folder):
        return False
    with os.scandir(folder) as entries:
        return any(entry.is_file() for entry in entries)


def get_unique_path(path, reserve=False):
    with _UNIQUE_PATH_LOCK:
        if not os.path.exists(path) and path not in _RESERVED_OUTPUT_PATHS:
            if reserve:
                _RESERVED_OUTPUT_PATHS.add(path)
            return path
        folder, filename = os.path.split(path)
        name, ext = os.path.splitext(filename)
        idx = 1
        while True:
            candidate = os.path.join(folder, f"{name}_dup{idx}{ext}")
            if not os.path.exists(candidate) and candidate not in _RESERVED_OUTPUT_PATHS:
                if reserve:
                    _RESERVED_OUTPUT_PATHS.add(candidate)
                return candidate
            idx += 1


def release_reserved_path(path):
    with _UNIQUE_PATH_LOCK:
        _RESERVED_OUTPUT_PATHS.discard(path)


def sanitize_csv_cell(value):
    text = "" if value is None else str(value)
    if not text:
        return text
    dangerous_prefixes = ("=", "+", "-", "@", "\t", "\r", "\n")
    if text.startswith(dangerous_prefixes):
        return "'" + text
    if text[:1].isspace() and text.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def validate_output_dir(source_dir, output_root):
    source_abs = os.path.realpath(source_dir)
    output_abs = os.path.realpath(output_root)
    if source_abs == output_abs:
        raise ValueError("输出目录不能与源目录相同。")
    try:
        common = os.path.commonpath([source_abs, output_abs])
    except ValueError:
        return
    if common == source_abs:
        raise ValueError("输出目录不能位于源目录内，请选择源目录外部路径。")
    if common == output_abs:
        raise ValueError("输出目录不能包含源目录，请选择源目录外部路径。")


def suggest_output_dir(source_dir):
    source_abs = os.path.realpath(source_dir)
    parent = os.path.dirname(source_abs.rstrip("\\/")) or source_abs
    name = os.path.basename(source_abs.rstrip("\\/")) or "输出"
    return os.path.join(parent, f"{name}_输出")


def unique_preserve_order(items):
    seen = set()
    unique = []
    duplicates = 0
    for item in items:
        if item in seen:
            duplicates += 1
            continue
        seen.add(item)
        unique.append(item)
    return unique, duplicates


def image_to_buffer(img, image_format, **save_kwargs):
    buffer = io.BytesIO()
    img.save(buffer, format=image_format, **save_kwargs)
    return buffer


def write_image_buffer(buffer, path):
    with open(path, 'wb') as f:
        f.write(buffer.getvalue())


def copy_to_manual_review(src_path, manual_dir):
    if not manual_dir or not src_path or not os.path.isfile(src_path):
        return None
    os.makedirs(manual_dir, exist_ok=True)
    dst_path = os.path.join(manual_dir, os.path.basename(src_path))
    if os.path.exists(dst_path):
        dst_path = get_unique_path(dst_path)
    shutil.copy2(src_path, dst_path)
    return dst_path


def remove_output_file_quiet(path):
    """从提取/处理输出目录移除不合格图片，避免后续被打包。"""
    if not path or not os.path.isfile(path):
        return False
    try:
        os.remove(path)
        return True
    except OSError:
        return False


def send_to_manual_and_remove_output(output_path, manual_dir, manual_source_lookup=None):
    """复制原始文件到手动处理目录，并从提取目录删除处理失败/不合格输出。"""
    manual_copy = None
    if manual_dir:
        manual_src = resolve_manual_source(output_path, manual_source_lookup)
        manual_copy = copy_to_manual_review(manual_src, manual_dir)
    removed = remove_output_file_quiet(output_path)
    return manual_copy, removed


def add_manual_source_aliases(lookup, current_name, src_path):
    for key in (current_name, os.path.splitext(current_name)[0]):
        lookup[os.path.normcase(key)] = src_path


def build_manual_source_lookup(source_dir, files, clean_detail_name=False):
    lookup = {}
    for fn in files:
        src_path = os.path.join(source_dir, fn)
        add_manual_source_aliases(lookup, fn, src_path)
        if clean_detail_name:
            add_manual_source_aliases(lookup, clean_detail_suffix(fn), src_path)
    return lookup


def resolve_manual_source(current_path, source_lookup=None):
    if source_lookup:
        current_name = os.path.basename(current_path)
        for key in (
            current_name,
            clean_detail_suffix(current_name),
            os.path.splitext(current_name)[0],
            os.path.splitext(clean_detail_suffix(current_name))[0],
        ):
            src_path = source_lookup.get(os.path.normcase(key))
            if src_path and os.path.isfile(src_path):
                return src_path
    return current_path


def make_temp_image_path(folder, basename):
    return os.path.join(folder, f".tmp_{threading.get_ident()}_{uuid.uuid4().hex}_{basename}")


def flatten_to_white_rgb(img):
    rgba = img.convert('RGBA')
    bg = Image.new('RGB', rgba.size, (255, 255, 255))
    bg.paste(rgba, mask=rgba.getchannel('A'))
    return bg


def copy_files_parallel(source_dir, out_dir, files, log, label, stop_event=None):
    os.makedirs(out_dir, exist_ok=True)
    total = len(files)
    if total == 0:
        return 0
    copied = 0
    failed = 0
    progress_step = max(1, total // PROGRESS_REPORT_FRACTION)

    def copy_one(fn):
        if stop_event and stop_event.is_set():
            return None
        src = os.path.join(source_dir, fn)
        dst = get_unique_path(os.path.join(out_dir, fn), reserve=True)
        try:
            shutil.copy2(src, dst)
        except Exception:
            release_reserved_path(dst)
            raise
        return fn

    workers = min(COPY_WORKERS, total)
    executor = ThreadPoolExecutor(max_workers=workers)
    futures = {}
    try:
        file_iter = iter(files)
        for _ in range(workers):
            try:
                fn = next(file_iter)
            except StopIteration:
                break
            futures[executor.submit(copy_one, fn)] = fn

        stopping = False
        while futures:
            if stop_event and stop_event.is_set() and not stopping:
                log.write(f"  [停止] 用户中止{label}复制")
                stopping = True
            done, _ = wait(list(futures), return_when=FIRST_COMPLETED, timeout=EXECUTOR_POLL_SECONDS)
            for future in done:
                fn = futures.pop(future, None)
                if fn is None:
                    continue
                try:
                    if future.result() is not None:
                        copied += 1
                except Exception as e:
                    failed += 1
                    log.write(f"  [失败] 复制 {fn or '(未知文件)'}: {e}")
                if copied and copied % progress_step == 0:
                    log.write(f"  {label}复制进度: {copied}/{total}")
                if not stopping and not (stop_event and stop_event.is_set()):
                    try:
                        next_fn = next(file_iter)
                        futures[executor.submit(copy_one, next_fn)] = next_fn
                    except StopIteration:
                        pass
    finally:
        executor.shutdown(wait=True, cancel_futures=True)

    if failed:
        log.write(f"  [警告] {label}复制失败 {failed} 张")
    return copied


def clean_detail_names_in_dir(detail_dir, log, stop_event=None, manual_source_lookup=None):
    if not os.path.isdir(detail_dir):
        return 0, {}
    renamed = 0
    output_lookup = {}
    for fn in list(iter_image_files(detail_dir)):
        if stop_event and stop_event.is_set():
            log.write("  [停止] 用户中止详情图改名")
            break
        new_name = clean_detail_suffix(fn)
        src_original = resolve_manual_source(os.path.join(detail_dir, fn), manual_source_lookup)
        if new_name == fn:
            add_manual_source_aliases(output_lookup, fn, src_original)
            continue
        src = os.path.join(detail_dir, fn)
        dst = get_unique_path(os.path.join(detail_dir, new_name))
        try:
            os.replace(src, dst)
            renamed += 1
            add_manual_source_aliases(output_lookup, os.path.basename(dst), src_original)
        except Exception as e:
            log.write(f"  [失败] 详情图改名 {fn}: {e}")
    if renamed:
        log.write(f"  [完成] 已去掉 {renamed} 张详情图文件名中的'_详情图'")
    return renamed, output_lookup




def step1_detail(source_dir, out_dir, log, stop_event=None, files=None):
    if files is None:
        files = [f for f in iter_image_files(source_dir) if '_详情图' in f]
    if not files:
        log.write("  [信息] 无含'_详情图'的图片，跳过")
        os.makedirs(out_dir, exist_ok=True)
        return out_dir
    c = copy_files_parallel(source_dir, out_dir, files, log, "详情图", stop_event)
    log.write(f"  [完成] 已复制 {c} 张详情图至 {out_dir}（源文件未改动）")
    return out_dir


def step2_main(source_dir, out_dir, log, stop_event=None, files=None):
    if files is None:
        files = [f for f in iter_image_files(source_dir) if '_详情图' not in f]
    if not files:
        log.write("  [信息] 无主图图片，跳过")
        os.makedirs(out_dir, exist_ok=True)
        return out_dir
    c = copy_files_parallel(source_dir, out_dir, files, log, "主图", stop_event)
    log.write(f"  [完成] 已复制 {c} 张主图至 {out_dir}（源文件未改动）")
    return out_dir


def step3_read_excel(excel_path, col, log, start_row=None, end_row=None, sheet_name=None, stop_event=None):
    if not os.path.isfile(excel_path):
        log.write(f"[错误] Excel文件不存在: {excel_path}")
        return []
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                log.write(f"[错误] 工作表 '{sheet_name}' 不存在，可用: {wb.sheetnames}")
                return []
            ws = wb[sheet_name]
        else:
            ws = wb.active
        col_idx = excel_col_to_index(col)
        auto_end_row = end_row is None
        start_row = start_row or 2
        if end_row is not None and end_row < start_row:
            log.write(f"[错误] 行范围无效: {start_row} ~ {end_row}")
            return []
        if end_row is None:
            log.write(f"  行范围: {start_row} ~ 自动")
        else:
            log.write(f"  行范围: {start_row} ~ {end_row}")
        barcodes = []
        if end_row is not None:
            total = end_row - start_row + 1
        else:
            max_row = ws.max_row or 0
            total = max_row - start_row + 1 if max_row >= start_row else None
        last_pct = -1
        empty_streak = 0
        stopped_early = False
        for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row,
                                             min_col=col_idx, max_col=col_idx, values_only=True)):
            if stop_event and stop_event.is_set():
                log.write("  [停止] 用户中止Excel读取")
                break
            val = row[0] if row else None
            if val is not None:
                if isinstance(val, float) and val.is_integer():
                    val = int(val)
                bc = str(val).strip()
                if bc:
                    barcodes.append(bc)
                    empty_streak = 0
                else:
                    empty_streak += 1
            else:
                empty_streak += 1
            if auto_end_row and empty_streak >= EXCEL_EMPTY_ROW_BREAK_THRESHOLD:
                stopped_early = True
                log.write(
                    f"  [信息] 连续 {EXCEL_EMPTY_ROW_BREAK_THRESHOLD} 行为空，"
                    "已提前结束读取"
                )
                break
            if total:
                pct = int((i + 1) / total * 100)
                if pct >= last_pct + 20:
                    last_pct = pct
                    log.write(f"  读取进度: {pct}% ({i + 1}/{total})")
            elif (i + 1) % 1000 == 0:
                log.write(f"  读取进度: 已扫描 {i + 1} 行")
        barcodes, dup = unique_preserve_order(barcodes)
        if stopped_early:
            log.write("  [信息] 如需读取空行后的数据，请手动填写结束行")
        log.write(f"  [完成] 读取 {len(barcodes)} 个条码（去重 {dup} 个）")
        return barcodes
    finally:
        wb.close()


def excel_col_to_index(col):
    col = (col or "").strip().upper()
    if not re.match(r'^[A-Z]+$', col):
        raise ValueError(f"非法列名: {col}")
    col_idx = 0
    for ch in col:
        col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
    return col_idx


def read_excel_preview(excel_path, sheet_name=None, col_text='A'):
    """读取 Excel 表头列表和当前列表头。供 UI 后台线程调用。"""
    if not excel_path or not os.path.isfile(excel_path):
        return [], None
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        headers = []
        max_col = ws.max_column or 0
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val is not None:
                label = str(val).strip()
                if label:
                    col_letter = ""
                    n = col_idx
                    while n > 0:
                        n, rem = divmod(n - 1, 26)
                        col_letter = chr(65 + rem) + col_letter
                    headers.append((col_letter, label))
        cell_val = None
        if col_text and re.match(r'^[A-Z]+$', col_text):
            cell_val = ws.cell(row=1, column=excel_col_to_index(col_text)).value
        return headers, cell_val
    finally:
        wb.close()


def _build_barcode_to_files(files, barcode_set, clean_detail_name=False):
    barcode_to_files = {}
    for fn in files:
        cn = os.path.splitext(clean_detail_suffix(fn) if clean_detail_name else fn)[0]
        cn_key = cn.upper()
        if cn_key in barcode_set:
            barcode_to_files.setdefault(cn_key, []).append(fn)
        parts = cn_key.split('_')
        for i in range(len(parts) - 1):
            candidate = '_'.join(parts[:i + 1])
            if candidate in barcode_set:
                barcode_to_files.setdefault(candidate, []).append(fn)
    return barcode_to_files


def step4_match_preview(files, barcodes, log, clean_detail_name=False, stop_event=None):
    all_files = list(files)
    if not all_files:
        log.write("  [信息] 预览源无图片，跳过")
        return
    barcode_set = {str(b).upper() for b in barcodes}
    barcode_to_files = _build_barcode_to_files(all_files, barcode_set, clean_detail_name=clean_detail_name)

    matched_total = 0
    preview_total = 0
    unmatched = []
    total = len(barcodes)
    last_pct = -1
    for idx, barcode in enumerate(barcodes):
        if stop_event and stop_event.is_set():
            log.write("  [停止] 用户中止匹配预览")
            break
        matched = barcode_to_files.get(str(barcode).upper(), [])
        if not matched:
            unmatched.append(barcode)
            continue
        matched_total += len(matched)
        for fn in matched:
            out_name = clean_detail_suffix(fn) if clean_detail_name else fn
            log.write(f"  [预览] {fn} → {out_name}")
            preview_total += 1
        pct = int((idx + 1) / total * 100)
        if pct >= last_pct + 20:
            last_pct = pct
            log.write(f"  匹配进度: {pct}% ({idx + 1}/{total})")
    log.write(f"  [结果] 匹配{matched_total}张, 预览{preview_total}张, 失败0张, 未匹配{len(unmatched)}个条码")


def step4_match(source_dir, output_dir, barcodes, log, clean_detail_name=False,
                dry_run=False, stop_event=None, copy_mode=False, manual_source_lookup=None):
    if not os.path.isdir(source_dir):
        log.write(f"  源文件夹不存在: {source_dir}，跳过")
        return {}
    if not dry_run:
        os.makedirs(output_dir, exist_ok=True)
    all_files = sorted(iter_image_files(source_dir))
    if not all_files:
        log.write("  [信息] 源文件夹无图片，跳过")
        return {}
    barcode_set = {str(b).upper() for b in barcodes}
    barcode_to_files = _build_barcode_to_files(all_files, barcode_set, clean_detail_name=clean_detail_name)

    matched_total = 0
    moved_total = 0
    failed_total = 0
    unmatched = []
    file_log = []
    output_lookup = {}
    total = len(barcodes)
    last_pct = -1
    processed_files = set()

    op_name = "复制" if copy_mode else "剪切"
    for idx, barcode in enumerate(barcodes):
        if stop_event and stop_event.is_set():
            log.write("  [停止] 用户中止匹配")
            break
        matched = barcode_to_files.get(str(barcode).upper(), [])
        if not matched:
            unmatched.append(barcode)
            continue
        matched_total += len(matched)
        for fn in matched:
            src = os.path.join(source_dir, fn)
            if not copy_mode and src in processed_files:
                continue
            out_name = clean_detail_suffix(fn) if clean_detail_name else fn
            dst = get_unique_path(os.path.join(output_dir, out_name))
            if dry_run:
                log.write(f"  [预览] {fn} → {out_name}")
                moved_total += 1
            else:
                try:
                    fsize = os.path.getsize(src)
                    if copy_mode:
                        shutil.copy2(src, dst)
                    else:
                        shutil.move(src, dst)
                    moved_total += 1
                    processed_files.add(src)
                    src_original = resolve_manual_source(src, manual_source_lookup)
                    add_manual_source_aliases(output_lookup, os.path.basename(dst), src_original)
                    file_log.append({'src': src, 'dst': dst, 'size': fsize,
                                     'barcode': barcode, 'status': 'ok'})
                except Exception as e:
                    failed_total += 1
                    log.write(f"  [失败] {fn}: {e}")
        pct = int((idx + 1) / total * 100)
        if pct >= last_pct + 20:
            last_pct = pct
            log.write(f"  匹配进度: {pct}% ({idx + 1}/{total})")
    log.write(f"  [结果] 匹配{matched_total}张, {op_name}{moved_total}张, "
              f"失败{failed_total}张, 未匹配{len(unmatched)}个条码")
    if file_log:
        log_path = os.path.join(output_dir, '匹配日志.csv')
        try:
            with open(log_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['原路径', '目标路径', '文件大小', '条码', '状态'])
                for item in file_log:
                    writer.writerow([
                        sanitize_csv_cell(item['src']),
                        sanitize_csv_cell(item['dst']),
                        item['size'],
                        sanitize_csv_cell(item['barcode']),
                        sanitize_csv_cell(item['status']),
                    ])
            log.write(f"  [日志] 详情已写入 {log_path}")
        except Exception as e:
            log.write(f"  [警告] 写入日志失败: {e}")
    return output_lookup


def _resize_stretch(img, target):
    return img.resize(target, RESAMPLE_LANCZOS)


def _resize_crop(img, target):
    tw, th = target
    w, h = img.size
    ratio = max(tw / w, th / h)
    new_w, new_h = int(w * ratio), int(h * ratio)
    img = img.resize((new_w, new_h), RESAMPLE_LANCZOS)
    left = (new_w - tw) // 2
    top = (new_h - th) // 2
    return img.crop((left, top, left + tw, top + th))


def _resize_fit(img, target, bg_color=(255, 255, 255)):
    tw, th = target
    w, h = img.size
    ratio = min(tw / w, th / h)
    new_w, new_h = int(w * ratio), int(h * ratio)
    img = img.resize((new_w, new_h), RESAMPLE_LANCZOS)
    canvas = Image.new('RGB', target, bg_color)
    x = (tw - new_w) // 2
    y = (th - new_h) // 2
    canvas.paste(img, (x, y))
    return canvas


def _process_main_image(main_dir, fn, resize_mode='crop', force_format=None, stop_event=None,
                        manual_dir=None, manual_source_lookup=None):
    resize_fns = {'stretch': _resize_stretch, 'crop': _resize_crop, 'fit': _resize_fit}
    resize_fn = resize_fns.get(resize_mode, _resize_crop)
    fp = os.path.join(main_dir, fn)
    name, ext = os.path.splitext(fn)
    ext_lower = ext.lower()
    tmp_path = None
    out_path = None
    unresolved_copy = None
    try:
        if stop_event and stop_event.is_set():
            return MainImageResult(fn, True, None, unresolved_copy)
        with Image.open(fp) as img:
            if stop_event and stop_event.is_set():
                return MainImageResult(fn, True, None, unresolved_copy)
            img = ImageOps.exif_transpose(img)
            w, h = img.size
            needs_resize = (w != TARGET_SIZE[0] or h != TARGET_SIZE[1])
            target_ext = f".{force_format}" if force_format else None
            is_standard = ext_lower in {'.jpg', '.jpeg', '.png'}
            needs_convert = not is_standard
            fsize = os.path.getsize(fp)
            needs_compress = fsize > MAIN_IMAGE_MAX_BYTES
            if not (needs_resize or needs_convert or needs_compress):
                return MainImageResult(fn, True, None, unresolved_copy)
            if needs_resize:
                img = resize_fn(img, TARGET_SIZE)
            if ext_lower in {'.jpg', '.jpeg'}:
                out_ext = ext_lower
                out_format = 'JPEG'
            elif ext_lower == '.png':
                out_ext = '.png'
                out_format = 'PNG'
            elif force_format:
                out_ext = target_ext
                out_format = 'JPEG' if force_format == 'jpg' else 'PNG'
            else:
                out_ext = '.jpg'
                out_format = 'JPEG'
            out_name = name + out_ext
            out_path = os.path.join(main_dir, out_name)
            if out_name != fn and os.path.exists(out_path):
                out_path = get_unique_path(out_path, reserve=True)
                out_name = os.path.basename(out_path)
            tmp_path = make_temp_image_path(main_dir, out_name)
            if out_format == 'JPEG' and img.mode in ('RGBA', 'P', 'LA'):
                img = flatten_to_white_rgb(img)
            if img.mode not in ('RGB', 'RGBA'):
                if out_format == 'JPEG':
                    img = img.convert('RGB')
                else:
                    img = img.convert('RGBA')
            if out_format == 'JPEG':
                best_buffer = image_to_buffer(img, 'JPEG', quality=MAIN_JPEG_QUALITY)
            else:
                best_buffer = image_to_buffer(img, 'PNG', optimize=True)
            if best_buffer.tell() > MAIN_IMAGE_MAX_BYTES:
                if out_format == 'JPEG':
                    quality = MAIN_JPEG_COMPRESS_START_QUALITY
                    while quality >= MAIN_JPEG_MIN_QUALITY:
                        if stop_event and stop_event.is_set():
                            return MainImageResult(fn, True, None, unresolved_copy)
                        if best_buffer.tell() <= MAIN_IMAGE_MAX_BYTES:
                            break
                        best_buffer = image_to_buffer(img, 'JPEG', quality=quality)
                        quality -= 10
                else:
                    if img.mode in ('RGBA', 'LA'):
                        p_img = img.quantize(colors=256, method=Image.Quantize.FASTOCTREE).convert('RGBA')
                    else:
                        p_img = img.convert('P', palette=Image.Palette.ADAPTIVE, colors=256)
                    best_buffer = image_to_buffer(p_img, 'PNG', optimize=True)
            write_image_buffer(best_buffer, tmp_path)
            os.replace(tmp_path, out_path)
            if os.path.normcase(out_path) != os.path.normcase(fp) and os.path.exists(fp):
                try:
                    os.remove(fp)
                except OSError as e:
                    print(f"  [警告] {fn}: 主图转换成功，但旧文件删除失败: {e}")
        if best_buffer.tell() > MAIN_IMAGE_MAX_BYTES:
            unresolved_copy, _ = send_to_manual_and_remove_output(
                out_path, manual_dir, manual_source_lookup
            )
        return MainImageResult(out_name, True, None, unresolved_copy)
    except Exception as e:
        if out_path:
            release_reserved_path(out_path)
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        if manual_dir:
            unresolved_copy, _ = send_to_manual_and_remove_output(
                fp, manual_dir, manual_source_lookup
            )
        return MainImageResult(fn, False, str(e), unresolved_copy)


def step6_process_main(main_dir, log, stop_event=None, resize_mode='crop', force_format=None,
                       manual_dir=None, manual_source_lookup=None):
    files = sorted(iter_image_files(main_dir))
    total = len(files)
    if total == 0:
        log.write("  [信息] 主图输出文件夹无图片，跳过")
        return 0
    progress_step = max(1, total // PROGRESS_REPORT_FRACTION)
    workers = min(IMAGE_WORKERS, total)
    processed = 0
    unresolved_count = 0
    executor = ThreadPoolExecutor(max_workers=workers)
    futures = {}
    try:
        file_iter = iter(files)
        for _ in range(workers):
            try:
                fn = next(file_iter)
            except StopIteration:
                break
            futures[executor.submit(_process_main_image, main_dir, fn, resize_mode, force_format,
                                    stop_event, manual_dir, manual_source_lookup)] = fn
        stopping = False
        while futures:
            if stop_event and stop_event.is_set() and not stopping:
                log.write("  [停止] 用户中止主图处理")
                stopping = True
            done, _ = wait(list(futures), return_when=FIRST_COMPLETED, timeout=EXECUTOR_POLL_SECONDS)
            for future in done:
                futures.pop(future, None)
                processed += 1
                result = future.result()
                if result.manual_copy:
                    unresolved_count += 1
                    log.write(
                        f"  [提醒] {result.filename} 已复制到手动处理，并已从提取文件夹移除，不参与打包"
                    )
                if not result.ok:
                    log.write(f"  [失败] {result.filename}: {result.error}")
                if processed % progress_step == 0 or processed == total:
                    log.write(f"  主图处理进度: {processed}/{total}")
                if not stopping and not (stop_event and stop_event.is_set()):
                    try:
                        next_fn = next(file_iter)
                        futures[executor.submit(_process_main_image, main_dir, next_fn, resize_mode,
                                                force_format, stop_event, manual_dir,
                                                manual_source_lookup)] = next_fn
                    except StopIteration:
                        pass
    finally:
        executor.shutdown(wait=True, cancel_futures=True)
    log.write("  [完成] 主图处理完毕")
    return unresolved_count


def _process_detail_image_impl(detail_dir, fn, force_format=None, stop_event=None):
    fp = os.path.join(detail_dir, fn)
    name, ext = os.path.splitext(fn)
    ext_lower = ext.lower()
    converted = 0
    compressed = 0
    info = None

    is_standard = ext_lower in {'.jpg', '.jpeg', '.png'}
    if is_standard:
        target_ext = ext_lower
    elif force_format:
        target_ext = '.' + force_format
    else:
        target_ext = '.jpg'
    target_fmt = 'JPEG' if target_ext in {'.jpg', '.jpeg'} else 'PNG'
    needs_convert = not is_standard
    if stop_event and stop_event.is_set():
        return DetailImageResult(fn, True, None, converted, compressed, info)
    if needs_convert:
        out_path = None
        tmp_path = None
        replace_done = False
        old_fn = fn
        try:
            with Image.open(fp) as img:
                if stop_event and stop_event.is_set():
                    return DetailImageResult(fn, True, None, converted, compressed, info)
                img = ImageOps.exif_transpose(img)
                out_path = get_unique_path(os.path.join(detail_dir, name + target_ext), reserve=True)
                tmp_path = make_temp_image_path(detail_dir, os.path.basename(out_path))
                if img.mode in ('RGBA', 'P', 'LA') and target_fmt == 'JPEG':
                    img = flatten_to_white_rgb(img)
                elif target_fmt == 'JPEG' and img.mode != 'RGB':
                    img = img.convert('RGB')
                elif target_fmt == 'PNG' and img.mode not in ('RGB', 'RGBA'):
                    img = img.convert('RGBA')
                img.save(tmp_path, format=target_fmt)
            os.replace(tmp_path, out_path)
            replace_done = True
            tmp_path = None
            converted = 1
            fp = out_path
            fn = os.path.basename(out_path)
            ext_lower = os.path.splitext(fn)[1].lower()
            if os.path.normcase(out_path) != os.path.normcase(os.path.join(detail_dir, old_fn)):
                old_path = os.path.join(detail_dir, old_fn)
                if os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except OSError as e:
                        info = f"{old_fn}: 转换成功，但旧文件删除失败: {e}"
        except Exception as e:
            if out_path and not replace_done:
                release_reserved_path(out_path)
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
            return DetailImageResult(fn, False, f"格式转换 {fn}: {e}", converted, compressed, info)

    tmp = None
    try:
        if os.path.getsize(fp) <= DETAIL_IMAGE_MAX_BYTES:
            return DetailImageResult(fn, True, None, converted, compressed, info)
        out_fmt = 'JPEG' if ext_lower in {'.jpg', '.jpeg'} else 'PNG'
        tmp = make_temp_image_path(detail_dir, fn)
        with Image.open(fp) as img:
            if stop_event and stop_event.is_set():
                return DetailImageResult(fn, True, None, converted, compressed, info)
            img = ImageOps.exif_transpose(img)
            if out_fmt == 'JPEG':
                best_buffer = image_to_buffer(
                    img, 'JPEG', quality=DETAIL_JPEG_OPTIMIZE_QUALITY, optimize=True
                )
            else:
                best_buffer = image_to_buffer(img, 'PNG', optimize=True)
            if best_buffer.tell() <= DETAIL_IMAGE_MAX_BYTES:
                write_image_buffer(best_buffer, tmp)
                os.replace(tmp, fp)
                return DetailImageResult(fn, True, None, converted, 1, info)

            if out_fmt == 'PNG' and img.mode != 'P':
                quantize_method = Image.Quantize.FASTOCTREE if img.mode in ('RGBA', 'LA') else Image.Quantize.MEDIANCUT
                p_img = img.quantize(colors=256, method=quantize_method)
                if img.mode in ('RGBA', 'LA'):
                    p_img = p_img.convert('RGBA')
                best_buffer = image_to_buffer(p_img, 'PNG', optimize=True)
                if best_buffer.tell() <= DETAIL_IMAGE_MAX_BYTES:
                    write_image_buffer(best_buffer, tmp)
                    os.replace(tmp, fp)
                    return DetailImageResult(fn, True, None, converted, 1, info)

            if out_fmt == 'JPEG':
                for q in DETAIL_JPEG_QUALITIES:
                    if stop_event and stop_event.is_set():
                        return DetailImageResult(fn, True, None, converted, compressed, info)
                    best_buffer = image_to_buffer(img, 'JPEG', quality=q, optimize=True)
                    if best_buffer.tell() <= DETAIL_IMAGE_MAX_BYTES:
                        break
                if best_buffer.tell() <= DETAIL_IMAGE_MAX_BYTES:
                    write_image_buffer(best_buffer, tmp)
                    os.replace(tmp, fp)
                    return DetailImageResult(fn, True, None, converted, 1, info)

            base_w, base_h = img.size
            for factor in DETAIL_SCALE_FACTORS:
                if stop_event and stop_event.is_set():
                    return DetailImageResult(fn, True, None, converted, compressed, info)
                r_img = img.resize((int(base_w * factor), int(base_h * factor)), RESAMPLE_LANCZOS)
                if out_fmt == 'JPEG':
                    best_buffer = image_to_buffer(r_img, 'JPEG', quality=MAIN_JPEG_QUALITY, optimize=True)
                else:
                    best_buffer = image_to_buffer(r_img, 'PNG', optimize=True)
                if best_buffer.tell() <= DETAIL_IMAGE_MAX_BYTES:
                    break
            if best_buffer.tell() <= DETAIL_IMAGE_MAX_BYTES:
                write_image_buffer(best_buffer, tmp)
                os.replace(tmp, fp)
                compressed = 1
            else:
                info = f"{fn}: 压缩后仍超5MB，保留原文件"
        return DetailImageResult(fn, True, None, converted, compressed, info)
    except Exception as e:
        if tmp and os.path.exists(tmp):
            try:
                os.remove(tmp)
            except OSError:
                pass
        return DetailImageResult(fn, False, f"压缩 {fn}: {e}", converted, compressed, info)


def _process_detail_image(detail_dir, fn, force_format=None, stop_event=None,
                          manual_dir=None, manual_source_lookup=None):
    if not manual_dir:
        return _process_detail_image_impl(detail_dir, fn, force_format, stop_event)

    unresolved_copy = None
    result = _process_detail_image_impl(
        detail_dir, fn, force_format, stop_event
    )
    out_path = os.path.join(detail_dir, result.filename)
    still_too_large = (
        result.ok and result.compressed == 0 and os.path.isfile(out_path)
        and os.path.getsize(out_path) > DETAIL_IMAGE_MAX_BYTES
    )
    if still_too_large:
        unresolved_copy, _ = send_to_manual_and_remove_output(
            out_path, manual_dir, manual_source_lookup
        )
    elif not result.ok:
        failed_path = os.path.join(detail_dir, result.filename or fn)
        unresolved_copy, _ = send_to_manual_and_remove_output(
            failed_path, manual_dir, manual_source_lookup
        )
    return result._replace(manual_copy=unresolved_copy)


def step7_process_detail(detail_dir, log, stop_event=None, force_format=None,
                         manual_dir=None, manual_source_lookup=None):
    files = sorted(iter_image_files(detail_dir))
    total = len(files)
    if total == 0:
        log.write("  [信息] 详情图输出文件夹无图片，跳过")
        return 0
    converted = 0
    compressed = 0
    processed = 0
    unresolved_count = 0
    progress_step = max(1, total // PROGRESS_REPORT_FRACTION)
    workers = min(IMAGE_WORKERS, total)
    executor = ThreadPoolExecutor(max_workers=workers)
    futures = {}
    try:
        file_iter = iter(files)
        for _ in range(workers):
            try:
                fn = next(file_iter)
            except StopIteration:
                break
            futures[executor.submit(_process_detail_image, detail_dir, fn, force_format, stop_event,
                                    manual_dir, manual_source_lookup)] = fn
        stopping = False
        while futures:
            if stop_event and stop_event.is_set() and not stopping:
                log.write("  [停止] 用户中止详情图处理")
                stopping = True
            done, _ = wait(list(futures), return_when=FIRST_COMPLETED, timeout=EXECUTOR_POLL_SECONDS)
            for future in done:
                futures.pop(future, None)
                processed += 1
                result = future.result()
                if result.manual_copy:
                    unresolved_count += 1
                    log.write(
                        f"  [提醒] {result.filename} 已复制到手动处理，并已从提取文件夹移除，不参与打包"
                    )
                converted += result.converted
                compressed += result.compressed
                if not result.ok:
                    log.write(f"  [失败] {result.error}")
                if result.info:
                    log.write(f"  [信息] {result.info}")
                if processed % progress_step == 0 or processed == total:
                    log.write(f"  详情图处理进度: {processed}/{total}")
                if not stopping and not (stop_event and stop_event.is_set()):
                    try:
                        next_fn = next(file_iter)
                        futures[executor.submit(_process_detail_image, detail_dir, next_fn, force_format,
                                                stop_event, manual_dir, manual_source_lookup)] = next_fn
                    except StopIteration:
                        pass
    finally:
        executor.shutdown(wait=True, cancel_futures=True)
    log.write(f"  [完成] 详情图处理: 格式转换 {converted} 张, 压缩 {compressed} 张")
    return unresolved_count


def step8_zip(source_dir, target_dir, max_bytes, log, stop_event=None):
    files_info = []
    for root, _, files in os.walk(source_dir):
        for f in sorted(files):
            # 只允许图片进入压缩包
            if os.path.splitext(f)[1].lower() not in IMAGE_EXTS:
                continue
            fp = os.path.join(root, f)
            try:
                files_info.append((fp, os.path.getsize(fp)))
            except OSError as e:
                log.write(f"  [警告] 跳过无法读取的文件 {fp}: {e}")
    if not files_info:
        log.write(f"  [信息] {source_dir} 为空，跳过打包")
        return
    folder_name = os.path.basename(source_dir)
    os.makedirs(target_dir, exist_ok=True)
    for fp, fsize in files_info:
        if fsize > max_bytes:
            log.write(
                f"  [警告] {os.path.basename(fp)} 大小 {fsize / 1024 / 1024:.2f}MB "
                f"超过分卷上限 {max_bytes}字节，将单独生成超限zip"
            )
    idx = 1
    cur_size = 0
    zf = None
    current_zip_path = None
    total_size = sum(sz for _, sz in files_info)
    processed = 0
    last_pct = -1
    try:
        for fp, fsize in files_info:
            if stop_event and stop_event.is_set():
                log.write("  [停止] 用户中止打包")
                return
            if zf is None or cur_size + fsize > max_bytes:
                if zf:
                    zf.close()
                    zf = None
                    current_zip_path = None
                current_zip_path = os.path.join(target_dir, f'{folder_name}_{idx:03d}.zip')
                zf = zipfile.ZipFile(current_zip_path, 'w', zipfile.ZIP_DEFLATED)
                cur_size = 0
                idx += 1
                log.write(f"  [打包] 创建: {os.path.basename(current_zip_path)}")
            try:
                arcname = os.path.relpath(fp, source_dir).replace('\\', '/')
                zf.write(fp, arcname)
            except OSError as e:
                log.write(f"  [失败] 打包 {fp}: {e}")
                continue
            cur_size += fsize
            processed += fsize
            pct = int(processed / total_size * 100)
            if pct >= last_pct + 20:
                last_pct = pct
                log.write(f"  打包进度: {pct}%")
        if zf:
            zf.close()
            zf = None
            current_zip_path = None
        total_zips = idx - 1
        log.write(f"  [完成] 打包完毕，共 {total_zips} 个zip包")
    finally:
        if zf:
            zf.close()
        if current_zip_path and os.path.exists(current_zip_path):
            try:
                os.remove(current_zip_path)
                log.write(f"  [清理] 已删除未完成压缩包: {os.path.basename(current_zip_path)}")
            except OSError as e:
                log.write(f"  [警告] 删除未完成压缩包失败: {e}")


def run_all(
    source_dir,
    output_root,
    mode=1,
    excel_path=None,
    col='A',
    sheet_name=None,
    start_row=None,
    end_row=None,
    force_format=None,
    resize_mode='crop',
    stop_event=None,
    copy_mode=False,
):
    """执行完整处理流程。"""
    if not os.path.isdir(source_dir):
        raise ValueError(f"源文件夹不存在: {source_dir}")

    validate_output_dir(source_dir, output_root)
    _RESERVED_OUTPUT_PATHS.clear()

    no_excel = mode == 4 or not excel_path
    dry_run = (mode == 2)
    do_compress = mode == 1
    do_zip = mode == 1

    if excel_path and not no_excel and not os.path.isfile(excel_path):
        raise ValueError(f"Excel文件不存在: {excel_path}")

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]
    if dry_run:
        task_root = output_root
        log_path = None
        manual_dir = None
    else:
        task_root = get_unique_path(os.path.join(output_root, f"任务_{ts}"), reserve=True)
        os.makedirs(task_root, exist_ok=True)
        log_path = os.path.join(task_root, f"处理日志_{ts}.txt")
        manual_dir = os.path.join(task_root, MANUAL_REVIEW_DIR_NAME)
    log = LogWriter(log_path)

    log.write("=" * 60)
    log.write("图片处理工具")
    log.write("=" * 60)
    log.write(f"[信息] 模式: {MODE_NAMES.get(mode, '未知')}")
    log.write(f"[信息] 源文件夹: {source_dir}")
    if dry_run:
        log.write("[信息] 输出目录: 预览模式不使用（不会写入任何文件）")
    else:
        log.write(f"[信息] 输出目录: {task_root}")
    if no_excel:
        log.write("[信息] Excel: 全量处理（不按条码匹配）")
    else:
        log.write(f"[信息] Excel: {excel_path}")
        log.write(f"[信息] 条码列: {col}")
    log.write("[信息] 分类步骤: 始终复制，源文件夹不改动")
    log.write(f"[信息] Excel匹配提取步骤: {'复制（保留分类结果）' if copy_mode else '剪切（移动到提取目录）'}")
    log.write(f"[信息] 输出格式: JPG/JPEG/PNG保持原格式；非标准格式转{force_format.upper() if force_format else 'JPG'}")
    log.write(f"[信息] 主图缩放: {RESIZE_MODE_MAP.get(resize_mode, resize_mode)}")
    source_detail_files, source_main_files = split_source_images(source_dir)
    main_manual_lookup = build_manual_source_lookup(source_dir, source_main_files)
    detail_manual_lookup = build_manual_source_lookup(source_dir, source_detail_files, clean_detail_name=True)
    if dry_run:
        preview_detail_files = source_detail_files
        preview_main_files = source_main_files

    try:
        # 步骤1
        detail_dir = os.path.join(task_root, "详情图")
        log.write(f"\n[步骤1] 复制含'_详情图'的图片至 {detail_dir}...（源文件未改动）")
        if dry_run:
            log.write("  [预览] 跳过文件移动")
            log.write(f"  [预览] 详情图候选: {len(preview_detail_files)} 张")
        else:
            detail_dir = step1_detail(
                source_dir, detail_dir, log, stop_event=stop_event, files=source_detail_files
            )

        if stop_event and stop_event.is_set():
            log.write("\n[停止] 用户中止")
            return

        # 步骤2
        main_dir = os.path.join(task_root, "主图")
        log.write(f"\n[步骤2] 复制主图至 {main_dir}...（源文件未改动）")
        if dry_run:
            log.write("  [预览] 跳过文件移动")
            log.write(f"  [预览] 主图候选: {len(preview_main_files)} 张")
        else:
            main_dir = step2_main(
                source_dir, main_dir, log, stop_event=stop_event, files=source_main_files
            )

        if stop_event and stop_event.is_set():
            log.write("\n[停止] 用户中止")
            return

        # 匹配
        if no_excel:
            detail_output = detail_dir
            main_output = main_dir
            log.write("\n[信息] 全量处理模式：已跳过条码读取和匹配")
            if not dry_run:
                log.write("\n[步骤3] 全量模式详情图改名...")
                _, detail_manual_lookup = clean_detail_names_in_dir(
                    detail_output, log, stop_event=stop_event,
                    manual_source_lookup=detail_manual_lookup
                )
        else:
            log.write("\n[步骤3] 读取Excel条码...")
            barcodes = step3_read_excel(excel_path, col, log, start_row, end_row, sheet_name, stop_event=stop_event)
            if not barcodes:
                log.write("[错误] 未读取到任何条码，程序退出")
                return

            if stop_event and stop_event.is_set():
                log.write("\n[停止] 用户中止")
                return

            detail_output = os.path.join(task_root, "详情图提取")
            log.write(f"\n[步骤4] 详情图条码匹配提取（{'复制' if copy_mode else '剪切'}） -> {detail_output}")
            if dry_run:
                step4_match_preview(preview_detail_files, barcodes, log,
                                    clean_detail_name=True, stop_event=stop_event)
            else:
                detail_manual_lookup = step4_match(
                    detail_dir, detail_output, barcodes, log,
                    clean_detail_name=True, dry_run=dry_run,
                    stop_event=stop_event, copy_mode=copy_mode,
                    manual_source_lookup=detail_manual_lookup
                )

            if stop_event and stop_event.is_set():
                log.write("\n[停止] 用户中止")
                return

            main_output = os.path.join(task_root, "主图提取")
            log.write(f"\n[步骤5] 主图条码匹配提取（{'复制' if copy_mode else '剪切'}） -> {main_output}")
            if dry_run:
                step4_match_preview(preview_main_files, barcodes, log,
                                    clean_detail_name=False, stop_event=stop_event)
            else:
                main_manual_lookup = step4_match(
                    main_dir, main_output, barcodes, log,
                    clean_detail_name=False, dry_run=dry_run,
                    stop_event=stop_event, copy_mode=copy_mode,
                    manual_source_lookup=main_manual_lookup
                )

            if stop_event and stop_event.is_set():
                log.write("\n[停止] 用户中止")
                return

        if dry_run:
            log.write("\n" + "=" * 60)
            log.write("  预览结束（未修改任何文件）")
            log.write("  确认无误后重新运行即可")
            log.write("=" * 60)
            return

        # 步骤6+7
        manual_count = 0
        if do_compress:
            log.write("\n[步骤6] 主图处理...")
            if dir_has_files(main_output):
                manual_count += step6_process_main(main_output, log, stop_event=stop_event,
                                                   resize_mode=resize_mode, force_format=force_format,
                                                   manual_dir=manual_dir,
                                                   manual_source_lookup=main_manual_lookup)
            else:
                log.write("  [信息] 主图输出文件夹为空，跳过")

            if stop_event and stop_event.is_set():
                log.write("\n[停止] 用户中止")
                return

            log.write("\n[步骤7] 详情图格式转换...")
            if dir_has_files(detail_output):
                manual_count += step7_process_detail(detail_output, log, stop_event=stop_event,
                                                     force_format=force_format, manual_dir=manual_dir,
                                                     manual_source_lookup=detail_manual_lookup)
            else:
                log.write("  [信息] 详情图输出文件夹为空，跳过")

            if stop_event and stop_event.is_set():
                log.write("\n[停止] 用户中止")
                return
        else:
            log.write("\n[步骤6/7] 跳过压缩处理")

        # 步骤8
        if manual_count:
            log.write(f"  [提醒] {manual_count} 张图片压缩后仍未达到大小要求，原图已复制到: {manual_dir}")

        if do_zip:
            log.write(f"\n[步骤8] 打包zip（每卷≤{ZIP_SPLIT_BYTES}字节）...")
            zip_dir = os.path.join(task_root, "压缩包")
            if dir_has_files(main_output):
                log.write("  --- 打包主图提取 ---")
                step8_zip(main_output, zip_dir, ZIP_SPLIT_BYTES, log, stop_event=stop_event)
            if dir_has_files(detail_output):
                log.write("  --- 打包详情图提取 ---")
                step8_zip(detail_output, zip_dir, ZIP_SPLIT_BYTES, log, stop_event=stop_event)

            if stop_event and stop_event.is_set():
                log.write("\n[停止] 用户中止")
                return
        else:
            log.write("\n[步骤8] 跳过打包")

        # 完成
        log.write("\n" + "=" * 60)
        log.write("  全部处理完成！")
        log.write(f"  主图输出: {main_output}")
        log.write(f"  详情图输出: {detail_output}")
        if do_zip:
            log.write(f"  zip包: {os.path.join(task_root, '压缩包')}")
        if log_path:
            log.write(f"  日志文件: {log_path}")
        if manual_count:
            log.write(f"  需要手动处理: {manual_dir}")
        log.write("=" * 60)

    except Exception as e:
        log.write(f"\n[错误] 处理异常: {e}")
        import traceback
        log.write(traceback.format_exc())
    finally:
        log.close()
        _RESERVED_OUTPUT_PATHS.clear()


# ============================================================
# UI 部分
# ============================================================

class EmittingStream(QObject):
    """将 print 输出重定向为 Qt 信号"""
    textWritten = Signal(str)

    def write(self, text):
        self.textWritten.emit(str(text))

    def flush(self):
        pass

    def isatty(self):
        return False


class DragLineEdit(QLineEdit):
    """支持把文件或文件夹拖到输入框里自动填路径。拖入时高亮边框提示。"""
    _accent_color = '#60A5FA'  # 默认值，由 _apply_app_style 覆盖

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet(
                f"background: palette(base); border: 2px solid {DragLineEdit._accent_color}; "
                "border-radius: 6px; padding: 0px 9px;"
            )
        else:
            super().dragEnterEvent(event)

    def dragLeaveEvent(self, event):
        self.setStyleSheet("")
        event.accept()

    def dropEvent(self, event):
        self.setStyleSheet("")
        urls = event.mimeData().urls()
        if urls:
            path = urls[0].toLocalFile()
            if path:
                self.setText(os.path.normpath(path))
                event.acceptProposedAction()
                return
        super().dropEvent(event)


class RoundComboBox(QComboBox):
    """支持圆角透明弹窗背景的下拉框。"""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._popup_chrome_applied = False

    def showPopup(self):
        # 1. 必须在系统绘制弹窗前设置透明属性，否则会导致全黑Bug
        if not self._popup_chrome_applied:
            view = self.view()
            if view:
                container = view.parentWidget()
                if container:
                    container.setObjectName("comboPopupContainer")
                    container.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
                    container.setWindowFlags(
                        Qt.WindowType.Popup
                        | Qt.WindowType.FramelessWindowHint
                        | Qt.WindowType.NoDropShadowWindowHint
                    )
                    # 用 ID 选择器只让容器自身透明，不能无选择器写样式
                    # 否则会级联覆盖子控件 QListView 的背景和边框
                    container.setStyleSheet(
                        "#comboPopupContainer { background: transparent; border: none; }"
                    )
            self._popup_chrome_applied = True
            
        # 2. 属性设置好之后，再调用系统的显示弹窗方法
        super().showPopup()


class AnimatedProgressBar(QProgressBar):
    """进度条，带微妙的流光效果。"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self._shimmer_x = -1.0
        self._timer = QTimer(self)
        self._timer.setInterval(SHIMMER_INTERVAL_MS)
        self._timer.timeout.connect(self._tick)

    def _tick(self):
        self._shimmer_x += SHIMMER_STEP
        if self._shimmer_x > 2.0:
            self._shimmer_x = -0.5
        self.update()

    def start_animation(self):
        self._shimmer_x = -0.5
        self._timer.start()

    def stop_animation(self):
        self._timer.stop()
        self._shimmer_x = -1.0
        self.update()

    def paintEvent(self, event):
        super().paintEvent(event)
        if self.maximum() == self.minimum() or self._shimmer_x < -0.5:
            return
        val = (self.value() - self.minimum()) / max(1, self.maximum() - self.minimum())
        if val <= 0:
            return
        fill_w = int(self.width() * val)
        if fill_w <= 0:
            return

        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        # 裁剪到填充区域（圆角与进度条外部一致）
        clip_path = QPainterPath()
        clip_path.addRoundedRect(QRectF(self.rect()), 8, 8)
        painter.setClipPath(clip_path)
        painter.setClipRect(0, 0, fill_w, self.height(), Qt.ClipOperation.IntersectClip)

        # 绘制流光条纹
        band_w = SHIMMER_BAND_WIDTH
        sx = fill_w * self._shimmer_x - band_w / 2
        grad = QLinearGradient(sx, 0, sx + band_w, 0)
        grad.setColorAt(0, QColor(255, 255, 255, 0))
        grad.setColorAt(0.5, QColor(255, 255, 255, 40))
        grad.setColorAt(1, QColor(255, 255, 255, 0))
        painter.fillRect(QRectF(sx, 0, band_w, self.height()), grad)
        painter.end()


class WorkerThread(QThread):
    """后台工作线程"""
    finished_signal = Signal()

    def __init__(self, kwargs):
        super().__init__()
        self.kwargs = kwargs

    def run(self):
        try:
            run_all(**self.kwargs)
        except Exception as e:
            print(f"\n[错误] {e}")
            import traceback
            print(traceback.format_exc())
        finally:
            self.finished_signal.emit()


class ExcelHeaderWorker(QThread):
    """后台读取 Excel 表头，避免大文件卡住 UI。"""
    loaded = Signal(str, str, str, list, object)
    failed = Signal(str, str, str, str)

    def __init__(self, excel_path, sheet_name, col_text):
        super().__init__()
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.col_text = col_text

    def run(self):
        try:
            headers, cell_val = read_excel_preview(self.excel_path, self.sheet_name, self.col_text)
            self.loaded.emit(self.excel_path, self.sheet_name or "", self.col_text, headers, cell_val)
        except Exception as e:
            self.failed.emit(self.excel_path, self.sheet_name or "", self.col_text, str(e))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("图片处理工具")
        self.resize(1000, 680)
        self._colors = self._theme_colors()

        self.stop_event = threading.Event()
        self.worker = None
        self.header_worker = None
        self.header_workers = []
        self._last_output_dir = None
        self._old_stdout = sys.stdout
        self._stdout_stream = None
        self._debounce_timer = QTimer(self)
        self._debounce_timer.setSingleShot(True)
        self._debounce_timer.setInterval(DEBOUNCE_MS)
        self._debounce_timer.timeout.connect(self._do_update_header_preview)

        self._setup_ui()
        self._update_excel_params_enabled()  # 初始没有 Excel

        self._stdout_stream = EmittingStream()
        self._stdout_stream.textWritten.connect(
            self.append_log, Qt.ConnectionType.QueuedConnection
        )
        sys.stdout = self._stdout_stream

    def _populate_col_combo(self, headers):
        """根据当前 Excel 路径重新填充条码列下拉框"""
        # 保存当前选中值
        old_col = self.cb_col.currentData()
        self.cb_col.blockSignals(True)
        self.cb_col.clear()
        if headers:
            idx_to_set = 0
            for i, (col_letter, header_text) in enumerate(headers):
                header_brief = (str(header_text) if header_text is not None else "").strip()
                if len(header_brief) > 10:
                    header_brief = header_brief[:10] + "..."
                display_text = f"{col_letter}-{header_brief}" if header_brief else col_letter
                self.cb_col.addItem(display_text, col_letter)
                self.cb_col.setItemData(i, f"{col_letter} - {header_text}", Qt.ItemDataRole.ToolTipRole)
                if col_letter == old_col:
                    idx_to_set = i
            self.cb_col.setCurrentIndex(idx_to_set)
        else:
            fallback = old_col or "A"
            self.cb_col.addItem(fallback, fallback)
            self.cb_col.setItemData(0, "未检测到表头，使用当前/默认列", Qt.ItemDataRole.ToolTipRole)
        self.cb_col.blockSignals(False)

    def _update_excel_params_enabled(self):
        """Excel 路径为空时禁用所有 Excel 参数控件"""
        has_excel = bool(self.le_excel.text().strip())
        self.cb_col.setEnabled(has_excel)
        self.le_sheet.setEnabled(has_excel)
        self.le_start.setEnabled(has_excel)
        self.le_end.setEnabled(has_excel)
        self._sync_row_range_input_style(has_excel)
        self.lbl_header_preview.setEnabled(has_excel)
        if not has_excel:
            self.lbl_header_preview.setText("表头：(未选择 Excel)")
            self.lbl_header_preview.setStyleSheet(f"color: {self._colors['muted']};")

    def _sync_row_range_input_style(self, enabled):
        """行范围输入框在嵌套布局里禁用态偶尔不吃全局 QSS，单独同步颜色。"""
        c = self._colors
        bg = c['input'] if enabled else c['card2']
        fg = c['text'] if enabled else c['muted']
        border = c['input_border'] if enabled else c['border']
        style = (
            f"background: {bg}; color: {fg}; border: 1px solid {border}; "
            "border-radius: 6px; min-height: 34px; max-height: 34px; padding: 0px 10px;"
        )
        self.le_start.setStyleSheet(style)
        self.le_end.setStyleSheet(style)

    def _theme_colors(self):
        """根据系统调色板生成浅色/暗色可读配色。"""
        app = QApplication.instance()
        is_dark = False
        if app:
            is_dark = app.palette().color(QPalette.ColorRole.Window).lightness() < 128
        if sys.platform == 'win32':
            try:
                import winreg
                with winreg.OpenKey(
                    winreg.HKEY_CURRENT_USER,
                    r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize"
                ) as key:
                    is_dark = winreg.QueryValueEx(key, "AppsUseLightTheme")[0] == 0
            except Exception:
                pass
        if is_dark:
            return {
                'window': '#111827',
                'card': '#1F2937',
                'card2': '#263244',
                'border': '#374151',
                'text': '#F3F4F6',
                'muted': '#AAB2BD',
                'input': '#111827',
                'input_border': '#6B7280',
                'accent': '#60A5FA',
                'accent_hover': '#3B82F6',
                'green': '#16A34A',
                'green_hover': '#15803D',
                'red': '#EF4444',
                'red_hover': '#DC2626',
                'log': '#0B1220',
                'help': '#152219',
                'tip': '#2A2112',
                'tip_text': '#FBBF24',
                'tip_border': '#D97706',
                'disabled_card': '#161E2A',
                'disabled_border': '#374151',
                'shadow': QColor(0, 0, 0, 45),
            }
        return {
            'window': '#F6F8FB',
            'card': '#FFFFFF',
            'card2': '#F8FAFD',
            'border': '#E4E8EF',
            'text': '#111827',
            'muted': '#6B7280',
            'input': '#FFFFFF',
            'input_border': '#D7DDE7',
            'accent': '#0078D4',
            'accent_hover': '#106EBE',
            'green': '#0EA348',
            'green_hover': '#098A37',
            'red': '#FF2D30',
            'red_hover': '#E52225',
            'log': '#FFFFFF',
            'help': '#F4FBF5',
            'tip': '#FFF8E8',
            'tip_text': '#B45309',
            'tip_border': '#F5C16C',
            'disabled_card': '#F0F3F8',
            'disabled_border': '#D6DCE7',
            'shadow': QColor(15, 23, 42, 18),
        }

    def _apply_app_style(self):
        c = self._colors
        # Qt QSS 不支持 data URI，需要把图标写到临时文件再引用
        # Checkbox 勾选图标
        checkbox_checked_svg = (
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="20" height="20">'
            f'<rect x="2" y="2" width="20" height="20" rx="6" ry="6" fill="{c["green"]}" />'
            f'<path d="M8 12 L11 15 L16 9" fill="none" stroke="#FFFFFF"'
            f' stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"/>'
            f'</svg>'
        )
        checkbox_unchecked_svg = (
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="20" height="20">'
            f'<rect x="2" y="2" width="20" height="20" rx="6" ry="6" '
            f'fill="{c["input"]}" stroke="{c["input_border"]}" stroke-width="2" />'
            f'</svg>'
        )
        cb_checked_ico = _write_temp_svg(checkbox_checked_svg, "cb_checked_ico_")
        cb_unchecked_ico = _write_temp_svg(checkbox_unchecked_svg, "cb_unchecked_ico_")
        checkbox_checked_css = f"image: url({cb_checked_ico});" if cb_checked_ico else f"background: {c['green']};"
        checkbox_unchecked_css = f"image: url({cb_unchecked_ico});" if cb_unchecked_ico else f"background: {c['input']};"

        # ComboBox 下拉箭头图标
        arrow_svg = (
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 12 8" width="12" height="8">'
            f'<path d="M1.5 1.5L6 6L10.5 1.5" fill="none" stroke="{c["muted"]}"'
            f' stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/>'
            f'</svg>'
        )
        arrow_ico = _write_temp_svg(arrow_svg, "cb_arrow_")

        # 开始按钮（播放）图标 - 启用态白色、禁用态灰色
        play_ico = _write_temp_svg(
            '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="20" height="20">'
            f'<polygon points="8,5 19,12 8,19" fill="white"/></svg>',
            "btn_play_"
        )
        play_ico_dis = _write_temp_svg(
            '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="20" height="20">'
            f'<polygon points="8,5 19,12 8,19" fill="{c["muted"]}"/></svg>',
            "btn_play_dis_"
        )

        # 停止按钮（方块）图标 - 启用态白色、禁用态灰色
        stop_ico = _write_temp_svg(
            '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="18" height="18">'
            '<rect x="6" y="6" width="12" height="12" rx="2" fill="white"/></svg>',
            "btn_stop_"
        )
        stop_ico_dis = _write_temp_svg(
            '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="18" height="18">'
            f'<rect x="6" y="6" width="12" height="12" rx="2" fill="{c["muted"]}"/></svg>',
            "btn_stop_dis_"
        )

        # 保存图标路径，_setup_ui 和启用/禁用切换时使用
        self._play_ico_path = play_ico
        self._stop_ico_path = stop_ico
        self._play_ico_dis_path = play_ico_dis
        self._stop_ico_dis_path = stop_ico_dis

        # 更新 DragLineEdit 的拖拽高亮颜色，跟随主题
        DragLineEdit._accent_color = c['accent']

        self.setStyleSheet(f"""
            QMainWindow, QWidget {{
                background: {c['window']};
                color: {c['text']};
            }}
            QFrame#CardFrame {{
                background: {c['card']};
                border: 1px solid {c['border']};
                border-radius: 8px;
            }}
            QLabel {{
                color: {c['text']};
                background: transparent;
            }}
            QLabel:disabled {{
                color: {c['muted']};
            }}
            QToolTip {{
                background: {c['card']};
                color: {c['text']};
                border: 1px solid {c['border']};
                border-radius: 4px;
                padding: 4px 8px;
            }}
            QLabel#sectionTitle {{
                color: {c['text']};
                background: transparent;
                font-size: 11pt;
                font-weight: 700;
                padding: 0 0 8px 0;
                border-bottom: 1px solid {c['border']};
            }}
            QLabel#hintLabel {{
                color: {c['muted']};
                font-size: 9pt;
            }}
            QLabel#fieldLabel {{
                font-size: 11pt;
                font-weight: 600;
            }}
            QLineEdit {{
                background: {c['input']};
                color: {c['text']};
                border: 2px solid {c['input_border']};
                border-radius: 6px;
                min-height: 34px;
                max-height: 34px;
                padding: 0px 9px;
            }}
            QComboBox {{
                background: {c['input']};
                color: {c['text']};
                border: 2px solid {c['input_border']};
                border-radius: 6px;
                min-height: 34px;
                max-height: 34px;
                padding: 0px 25px 0px 9px;
            }}
            QLineEdit:focus, QComboBox:focus {{
                border-color: {c['accent']};
            }}
            QLineEdit:hover, QComboBox:hover {{
                border-color: {c['accent']};
            }}
            QLineEdit:disabled, QComboBox:disabled {{
                color: {c['muted']};
                background: {c['card2']};
                border-color: {c['border']};
            }}
            QComboBox::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: top right;
                right: 10px;
                width: 14px;
                border: none;
                background: transparent;
            }}
            QComboBox::down-arrow {{
                image: url({arrow_ico});
                width: 10px;
                height: 7px;
            }}
            QListView#roundComboView {{
                background: {c['input']};
                color: {c['text']};
                border: 1px solid {c['input_border']};
                border-radius: 8px;
                outline: none;
                padding: 4px;
            }}
            QListView#roundComboView::item {{
                min-height: 26px;
                margin: 2px 4px;
                padding: 2px 8px;
                border-radius: 6px;
            }}
            QListView#roundComboView::item:selected {{
                background: {c['accent']};
                color: #FFFFFF;
            }}
            QPushButton {{
                background: {c['card2']};
                color: {c['text']};
                border: 2px solid {c['border']};
                border-radius: 7px;
                min-height: 34px;
                max-height: 34px;
                padding: 0px 11px;
            }}
            QPushButton:hover {{
                border-color: {c['accent']};
                background: {c['input']};
            }}
            QPushButton:pressed {{
                background: {c['border']};
                border-color: {c['muted']};
            }}
            QPushButton:disabled {{
                color: {c['muted']};
                background: {c['card2']};
                border-color: {c['border']};
            }}
            QPushButton#startButton {{
                background: {c['green']};
                border-color: {c['green_hover']};
                color: white;
                font-size: 13pt;
                font-weight: 800;
                min-height: 46px;
                max-height: 46px;
                border-radius: 8px;
                padding: 0px 16px;
            }}
            QPushButton#startButton:hover {{
                background: {c['green_hover']};
            }}
            QPushButton#startButton:pressed {{
                background: {_darken(c['green_hover'], 0.8)};
            }}
            QPushButton#startButton:disabled {{
                background: {c['card2']};
                border-color: {c['border']};
                color: {c['muted']};
            }}
            QPushButton#stopButton {{
                background: {c['red']};
                border-color: {c['red_hover']};
                color: white;
                font-size: 13pt;
                font-weight: 800;
                min-height: 46px;
                max-height: 46px;
                border-radius: 8px;
                padding: 0px 16px;
            }}
            QPushButton#stopButton:hover {{
                background: {c['red_hover']};
            }}
            QPushButton#stopButton:pressed {{
                background: {_darken(c['red_hover'], 0.8)};
            }}
            QPushButton#stopButton:disabled {{
                background: {c['card2']};
                border-color: {c['border']};
                color: {c['muted']};
            }}
            QPushButton#logToolButton {{
                min-height: 34px;
                max-height: 34px;
                padding: 0px 12px;
            }}
            QTextEdit {{
                background: {c['log']};
                color: {c['text']};
                border: 1px solid {c['border']};
                border-radius: 8px;
                padding: 10px;
                font-family: Consolas, "Courier New", monospace;
                font-size: 10pt;
            }}
            QTextEdit#helpText {{
                background: transparent;
                border: none;
            }}
            QFrame#helpContainer {{
                background: {c['card2']};
                border: 1px solid {c['border']};
                border-radius: 8px;
            }}
            QTabWidget#helpTabs::pane {{
                border: none;
                background: transparent;
                margin-top: 0px;
            }}
            QTabWidget#helpTabs QWidget {{
                background: transparent;
            }}
            QTabWidget#helpTabs QTabBar {{
                background: transparent;
            }}
            QTabWidget#helpTabs QTabBar::tab {{
                background: {c['card2']};
                border: 1px solid {c['border']};
                color: {c['muted']};
                padding: 7px 12px;
                min-width: 84px;
                font-weight: 600;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                border-bottom-left-radius: 0px;
                border-bottom-right-radius: 0px;
                margin-right: 0px;
                margin-bottom: 0px;
            }}
            QTabWidget#helpTabs QTabBar::tab:selected {{
                background: {c['card']};
                color: {c['text']};
                border-color: {c['input_border']};
                border-bottom-color: {c['card']};
                margin-bottom: -1px;
            }}
            QFrame#helpTabCard {{
                background: {c['card']};
                border: 1px solid {c['input_border']};
                border-top: none;
                border-top-left-radius: 0px;
                border-top-right-radius: 0px;
                border-bottom-left-radius: 8px;
                border-bottom-right-radius: 8px;
            }}
            QTextEdit#helpTabText {{
                background: {c['card']};
                border: none;
                padding: 10px;
            }}
            QFrame#modeHelpBox {{
                background: {c['help']};
                border: 1px solid {c['border']};
                border-radius: 8px;
            }}
            QLabel#modeHelpText {{
                color: {c['green']};
                background: transparent;
            }}
            QFrame#tipBox {{
                background: {c['tip']};
                border: 1px solid {c['tip_border']};
                border-radius: 8px;
            }}
            QLabel#tipText {{
                color: {c['tip_text']};
                background: transparent;
            }}
            QProgressBar {{
                background: {c['border']};
                color: {c['text']};
                border: none;
                border-radius: 8px;
                min-height: 20px;
                text-align: center;
                font-weight: 700;
            }}
            QProgressBar::chunk {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {c['green']}, stop:1 {_darken(c['green'], 0.78)});
                border-radius: 8px;
            }}
            QCheckBox {{
                color: {c['text']};
                background: transparent;
                spacing: 8px;
                min-height: 40px;
                max-height: 40px;
            }}
            QCheckBox::indicator {{
                width: 20px;
                height: 20px;
                border: none;
                background: transparent;
            }}
            QCheckBox::indicator:unchecked {{
                {checkbox_unchecked_css}
            }}
            QCheckBox::indicator:checked {{
                {checkbox_checked_css}
            }}
            QScrollBar:vertical {{
                background: transparent;
                width: 10px;
                margin: 4px 2px 4px 2px;
            }}
            QScrollBar::handle:vertical {{
                background: {c['input_border']};
                border-radius: 5px;
                min-height: 28px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {c['muted']};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
            QScrollBar:horizontal {{
                background: transparent;
                height: 10px;
                margin: 2px 4px 2px 4px;
            }}
            QScrollBar::handle:horizontal {{
                background: {c['input_border']};
                border-radius: 5px;
                min-width: 28px;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: {c['muted']};
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
            }}
        """)

    def _add_shadow(self, widget, blur=14, y=3):
        effect = QGraphicsDropShadowEffect(widget)
        effect.setBlurRadius(blur)
        effect.setOffset(0, y)
        effect.setColor(self._colors['shadow'])
        widget.setGraphicsEffect(effect)

    def _section_title(self, icon, text, color=None):
        label = QLabel(text if not icon else f"{icon}  {text}")
        label.setObjectName("sectionTitle")
        if color:
            label.setStyleSheet(f"QLabel#sectionTitle {{ color: {color}; }}")
        return label

    def _hint_label(self, text):
        label = QLabel(text)
        label.setObjectName("hintLabel")
        label.setWordWrap(True)
        return label

    def _field_label(self, text):
        label = QLabel(text)
        label.setObjectName("fieldLabel")
        return label

    def _setup_combo_view(self, combo):
        view = QListView()
        view.setObjectName("roundComboView")
        combo.setView(view)
        combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        combo.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon)
        combo.setMinimumContentsLength(12)
        return combo

    def _setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        ui_gap = 14
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(ui_gap, ui_gap, ui_gap, ui_gap)
        main_layout.setSpacing(ui_gap)

        top_layout = QHBoxLayout()
        top_layout.setSpacing(ui_gap)

        path_group = QFrame()
        path_group.setObjectName("CardFrame")
        path_group.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        path_group.setMinimumWidth(480)
        path_group.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self._add_shadow(path_group)
        path_layout = QGridLayout(path_group)
        path_layout.setHorizontalSpacing(10)
        path_layout.setVerticalSpacing(10)
        path_layout.setContentsMargins(14, 12, 14, 12)
        path_layout.setColumnMinimumWidth(0, 76)
        path_layout.setColumnMinimumWidth(1, 240)
        path_layout.setColumnMinimumWidth(2, 80)
        path_layout.setColumnStretch(1, 1)
        path_layout.addWidget(
            self._section_title("", "路径与文件设置"),
            0, 0, 1, 3,
            Qt.AlignmentFlag.AlignTop,
        )

        self.le_source = DragLineEdit()
        self.le_source.setMinimumWidth(240)
        self.le_source.setPlaceholderText(r"D:\ecommerce\images\source")
        self.le_source.setToolTip("原始图片所在文件夹。源文件不会在分类步骤中被改动。")
        btn_source = QPushButton("浏览...")
        btn_source.setFixedWidth(80)
        btn_source.setToolTip("选择原始图片所在文件夹。")
        btn_source.clicked.connect(lambda: self._browse_dir(self.le_source, is_source=True))
        path_layout.addWidget(self._field_label("源文件夹"), 1, 0, Qt.AlignmentFlag.AlignVCenter)
        path_layout.addWidget(self.le_source, 1, 1, Qt.AlignmentFlag.AlignVCenter)
        path_layout.addWidget(btn_source, 1, 2, Qt.AlignmentFlag.AlignVCenter)

        self.le_output = DragLineEdit()
        self.le_output.setMinimumWidth(240)
        self.le_output.setPlaceholderText(r"D:\ecommerce\images\output")
        self.le_output.setToolTip("处理结果保存目录。建议选择空文件夹或新文件夹，方便查看结果。")
        btn_output = QPushButton("浏览...")
        btn_output.setFixedWidth(80)
        btn_output.setToolTip("选择处理结果保存目录。")
        btn_output.clicked.connect(lambda: self._browse_dir(self.le_output, is_source=False))
        path_layout.addWidget(self._field_label("输出目录"), 2, 0, Qt.AlignmentFlag.AlignVCenter)
        path_layout.addWidget(self.le_output, 2, 1, Qt.AlignmentFlag.AlignVCenter)
        path_layout.addWidget(btn_output, 2, 2, Qt.AlignmentFlag.AlignVCenter)

        self.le_excel = DragLineEdit()
        self.le_excel.setMinimumWidth(240)
        self.le_excel.setPlaceholderText(r"D:\ecommerce\images\list\商品图清单.xlsx")
        self.le_excel.setToolTip("可选。选择 Excel 后会按条码列匹配图片；留空则不按条码匹配。")
        self.le_excel.textChanged.connect(self._update_excel_params_enabled)
        btn_excel = QPushButton("浏览...")
        btn_excel.setFixedWidth(80)
        btn_excel.setToolTip("选择包含条码清单的 Excel 文件（.xlsx 或 .xlsm）。")
        btn_excel.clicked.connect(self._browse_excel)
        path_layout.addWidget(self._field_label("Excel清单"), 3, 0, Qt.AlignmentFlag.AlignVCenter)
        path_layout.addWidget(self.le_excel, 3, 1, Qt.AlignmentFlag.AlignVCenter)
        path_layout.addWidget(btn_excel, 3, 2, Qt.AlignmentFlag.AlignVCenter)

        path_hint = self._hint_label("Excel 中需包含条码列，用于与图片文件名匹配；不选择 Excel 时会全量处理所有图片")
        path_hint.setContentsMargins(0, 2, 0, 0)
        path_layout.addWidget(path_hint, 4, 1, 1, 2, Qt.AlignmentFlag.AlignTop)
        top_layout.addWidget(path_group, 2)

        self.excel_group = QFrame()
        self.excel_group.setObjectName("CardFrame")
        self.excel_group.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.excel_group.setMinimumWidth(280)
        self.excel_group.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Preferred)
        self._add_shadow(self.excel_group)
        excel_layout = QGridLayout(self.excel_group)
        excel_layout.setHorizontalSpacing(10)
        excel_layout.setVerticalSpacing(12)
        excel_layout.setContentsMargins(14, 14, 14, 14)
        excel_layout.setColumnMinimumWidth(0, 54)
        excel_layout.setColumnStretch(1, 1)
        excel_layout.addWidget(
            self._section_title("", "Excel 参数配置"),
            0, 0, 1, 4,
            Qt.AlignmentFlag.AlignTop,
        )

        excel_layout.addWidget(self._field_label("条码列"), 1, 0)
        barcode_row = QHBoxLayout()
        barcode_row.setSpacing(10)
        barcode_row.setContentsMargins(0, 0, 0, 0)
        self.cb_col = RoundComboBox()
        self.cb_col.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.cb_col.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon)
        self.cb_col.setMinimumContentsLength(8)
        self._setup_combo_view(self.cb_col)
        self.cb_col.setToolTip("Excel 第一行会作为表头预览；请选择真正的条码列。")
        barcode_row.addWidget(self.cb_col, 1)
        self.lbl_header_preview = QLabel("表头：(未加载)")
        self.lbl_header_preview.setToolTip("显示当前条码列在 Excel 第一行对应的表头，便于检查是否选错列。")
        self.lbl_header_preview.setStyleSheet(f"color: {self._colors['muted']}; font-style: italic;")
        excel_layout.addLayout(barcode_row, 1, 1, 1, 3)
        excel_layout.addWidget(self.lbl_header_preview, 2, 1, 1, 3, Qt.AlignmentFlag.AlignLeft)

        excel_layout.addWidget(self._field_label("工作表"), 3, 0)
        self.le_sheet = QLineEdit()
        self.le_sheet.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.le_sheet.setPlaceholderText("Sheet1")
        self.le_sheet.setToolTip("可选。填写指定工作表名称；留空默认读取第一个工作表。")
        excel_layout.addWidget(self.le_sheet, 3, 1, 1, 3)

        excel_layout.addWidget(self._field_label("行范围"), 4, 0)
        self.row_range_wrap = QWidget()
        self.row_range_wrap.setStyleSheet("background: transparent;")
        self.row_range_wrap.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        row_range_layout = QHBoxLayout(self.row_range_wrap)
        row_range_layout.setSpacing(10)
        row_range_layout.setContentsMargins(0, 0, 0, 0)
        self.le_start = QLineEdit()
        self.le_start.setObjectName("rowRangeInput")
        self.le_start.setMinimumWidth(60)
        self.le_start.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.le_start.setPlaceholderText("2")
        self.le_start.setToolTip("可选。Excel 起始行号；留空默认从第 2 行开始，通常第 1 行是表头。")
        row_range_layout.addWidget(self.le_start, 1)
        self.lbl_to = self._field_label("至")
        self.lbl_to.setFixedWidth(30)
        self.lbl_to.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_to.setStyleSheet("background: transparent;")
        row_range_layout.addWidget(self.lbl_to)
        self.le_end = QLineEdit()
        self.le_end.setObjectName("rowRangeInput")
        self.le_end.setMinimumWidth(60)
        self.le_end.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.le_end.setPlaceholderText("999999")
        self.le_end.setToolTip("可选。Excel 结束行号；留空默认读取到最后一行。")
        row_range_layout.addWidget(self.le_end, 1)
        excel_layout.addWidget(self.row_range_wrap, 4, 1, 1, 3)

        excel_hint = self._hint_label("从第 2 行开始读取数据（通常第 1 行为表头）")
        excel_layout.addWidget(excel_hint, 5, 1, 1, 3)

        self.le_excel.textChanged.connect(self._debounce_header_preview)
        self.le_excel.textChanged.connect(self._update_excel_params_enabled)
        self.le_sheet.textChanged.connect(self._debounce_header_preview)
        self.cb_col.currentIndexChanged.connect(self._debounce_header_preview)
        top_layout.addWidget(self.excel_group, 1)

        opt_group = QFrame()
        opt_group.setObjectName("CardFrame")
        opt_group.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        opt_group.setMinimumWidth(320)
        opt_group.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Preferred)
        self._add_shadow(opt_group)
        opt_layout = QGridLayout(opt_group)
        opt_layout.setHorizontalSpacing(10)
        opt_layout.setVerticalSpacing(12)
        opt_layout.setContentsMargins(14, 14, 14, 14)
        opt_layout.setColumnMinimumWidth(0, 104)
        opt_layout.setColumnMinimumWidth(1, 170)
        opt_layout.setColumnStretch(1, 1)
        opt_layout.addWidget(
            self._section_title("", "处理选项"),
            0, 0, 1, 2,
            Qt.AlignmentFlag.AlignTop,
        )

        opt_layout.addWidget(self._field_label("处理流程"), 1, 0)
        self.cb_mode = RoundComboBox()
        self._setup_combo_view(self.cb_mode)
        self.cb_mode.addItems(["完整流程", "预览模式", "仅分类(无Excel)"])
        self.cb_mode.setToolTip("完整流程适合正式处理；预览模式适合检查匹配；仅分类适合整理图片。")
        self.cb_mode.setItemData(0, MODE_HELP_TEXT["完整流程"], Qt.ItemDataRole.ToolTipRole)
        self.cb_mode.setItemData(1, MODE_HELP_TEXT["预览模式"], Qt.ItemDataRole.ToolTipRole)
        self.cb_mode.setItemData(2, MODE_HELP_TEXT["仅分类(无Excel)"], Qt.ItemDataRole.ToolTipRole)
        opt_layout.addWidget(self.cb_mode, 1, 1)

        opt_layout.addWidget(self._field_label("裁剪方式"), 2, 0)
        self.cb_resize = RoundComboBox()
        self._setup_combo_view(self.cb_resize)
        self.cb_resize.addItems(["crop - 铺满裁剪", "stretch - 拉伸", "fit - 留白缩放"])
        self.cb_resize.setToolTip("选择主图缩放到 800×800 的方式。")
        self.cb_resize.setItemData(0, "铺满目标尺寸并裁剪超出部分（默认，适合方形展示图）", Qt.ItemDataRole.ToolTipRole)
        self.cb_resize.setItemData(1, "直接拉伸/压缩到目标尺寸（图片可能变形）", Qt.ItemDataRole.ToolTipRole)
        self.cb_resize.setItemData(2, "保持原始比例缩放，空白处填充白色（适合带留白的商品图）", Qt.ItemDataRole.ToolTipRole)
        opt_layout.addWidget(self.cb_resize, 2, 1)

        opt_layout.addWidget(self._field_label("输出格式"), 3, 0)
        self.cb_force_format = RoundComboBox()
        self._setup_combo_view(self.cb_force_format)
        self.cb_force_format.addItems(["保持原格式", "JPG", "PNG"])
        self.cb_force_format.setCurrentIndex(1)
        self.cb_force_format.setToolTip("JPG/JPEG/PNG 始终保持原格式；此选项只决定 WEBP/AVIF/BMP 等非标准格式转成 JPG 还是 PNG。")
        self.cb_force_format.setItemData(0, "保持原格式：JPG/JPEG/PNG 不转换格式；WEBP/AVIF/BMP 等非标准格式转 JPG", Qt.ItemDataRole.ToolTipRole)
        self.cb_force_format.setItemData(1, "仅将 WEBP/AVIF/BMP 等非标准格式转换为 JPG；JPG/JPEG/PNG 不互转", Qt.ItemDataRole.ToolTipRole)
        self.cb_force_format.setItemData(2, "仅将 WEBP/AVIF/BMP 等非标准格式转换为 PNG；JPG/JPEG/PNG 不互转", Qt.ItemDataRole.ToolTipRole)
        opt_layout.addWidget(self.cb_force_format, 3, 1)

        opt_layout.addWidget(self._field_label("文件处理方式"), 4, 0)
        self.cb_move = RoundComboBox()
        self._setup_combo_view(self.cb_move)
        self.cb_move.addItems(["剪切（移动源文件）", "复制（保留源文件）"])
        self.cb_move.setCurrentIndex(1)
        self.cb_move.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.cb_move.setToolTip("复制更安全，会保留分类目录中的原文件；剪切更省空间。")
        self.cb_move.setItemData(0, "匹配到的文件从分类目录移到提取目录（节省空间）", Qt.ItemDataRole.ToolTipRole)
        self.cb_move.setItemData(1, "匹配到的文件复制到提取目录，分类目录保留原文件（安全）", Qt.ItemDataRole.ToolTipRole)
        opt_layout.addWidget(self.cb_move, 4, 1)

        self.chk_open_output = QCheckBox("处理完成后打开输出目录")
        self.chk_open_output.setChecked(True)
        self.chk_open_output.setToolTip("处理完成后自动打开输出目录，方便查看结果。")
        opt_layout.addWidget(self.chk_open_output, 5, 0, 1, 2)
        top_layout.addWidget(opt_group, 1)
        main_layout.addLayout(top_layout)

        middle_layout = QHBoxLayout()
        middle_layout.setSpacing(ui_gap)

        help_group = QFrame()
        help_group.setObjectName("CardFrame")
        help_group.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self._add_shadow(help_group)
        help_layout = QVBoxLayout(help_group)
        help_layout.setSpacing(12)
        help_layout.setContentsMargins(16, 16, 16, 16)
        help_layout.addWidget(self._section_title("", "使用说明"))

        help_content_layout = QVBoxLayout()
        help_content_layout.setContentsMargins(0, 0, 0, 0)
        help_content_layout.setSpacing(12)

        self.help_tabs = QTabWidget()
        self.help_tabs.setObjectName("helpTabs")
        self.help_tabs.setDocumentMode(True)
        self.help_tabs.setUsesScrollButtons(False)
        self.help_tabs.tabBar().setDrawBase(False)
        self.help_tabs.tabBar().setExpanding(True)
        section_pattern = re.compile(r'【([^】]+)】\n(.*?)(?=\n【[^】]+】|$)', re.S)
        help_sections = section_pattern.findall(UI_HELP_TEXT)
        if not help_sections:
            help_sections = [("功能说明", UI_HELP_TEXT)]

        for title, body in help_sections:
            tab_page = QWidget()
            tab_layout = QVBoxLayout(tab_page)
            tab_layout.setContentsMargins(0, 0, 0, 0)
            tab_layout.setSpacing(0)

            tab_card = QFrame()
            tab_card.setObjectName("helpTabCard")
            tab_card.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
            tab_card_layout = QVBoxLayout(tab_card)
            tab_card_layout.setContentsMargins(10, 10, 10, 10)

            help_text = QTextEdit()
            help_text.setObjectName("helpTabText")
            help_text.setReadOnly(True)
            help_text.setPlainText(body.strip())
            help_text.setMinimumHeight(150)
            help_text.setToolTip("快速了解软件怎么用、各个选项的区别，以及输出目录里会生成什么。")
            tab_card_layout.addWidget(help_text)

            tab_layout.addWidget(tab_card)
            self.help_tabs.addTab(tab_page, title.strip())

        if self.help_tabs.count() > 0:
            first_help_text = self.help_tabs.widget(0).findChild(QTextEdit, "helpTabText")
            if first_help_text is not None:
                self.text_help = first_help_text
        help_content_layout.addWidget(self.help_tabs, 1)

        mode_help_box = QFrame()
        mode_help_box.setObjectName("modeHelpBox")
        mode_help_box.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        mode_help_layout = QVBoxLayout(mode_help_box)
        mode_help_layout.setContentsMargins(12, 10, 12, 10)
        mode_help_layout.setSpacing(0)
        self.lbl_mode_help = QLabel()
        self.lbl_mode_help.setObjectName("modeHelpText")
        self.lbl_mode_help.setWordWrap(True)
        self.lbl_mode_help.setToolTip("这里会随处理模式变化，说明当前模式实际会做什么。")
        mode_help_layout.addWidget(self.lbl_mode_help)
        self.cb_mode.currentTextChanged.connect(self._update_mode_help)
        self._update_mode_help(self.cb_mode.currentText())
        help_content_layout.addWidget(mode_help_box)

        tip_box = QFrame()
        tip_box.setObjectName("tipBox")
        tip_box.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        tip_layout = QVBoxLayout(tip_box)
        tip_layout.setContentsMargins(12, 10, 12, 10)
        tip_layout.setSpacing(0)
        tip_label = QLabel("提示：建议先小范围测试（如行范围 2-100），确认无误后再处理全部数据。")
        tip_label.setObjectName("tipText")
        tip_label.setWordWrap(True)
        tip_layout.addWidget(tip_label)
        help_content_layout.addWidget(tip_box)
        help_layout.addLayout(help_content_layout, 1)
        middle_layout.addWidget(help_group, 5)

        right_layout = QVBoxLayout()
        right_layout.setSpacing(12)
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(12)
        btn_layout.setContentsMargins(0, 0, 0, 0)
        self.btn_start = QPushButton("开始处理")
        self.btn_start.setObjectName("startButton")
        self.btn_start.setToolTip("按当前设置开始处理。处理期间下方日志会显示进度和结果。")
        self.btn_start.clicked.connect(self.start_process)
        self.btn_stop = QPushButton("停止处理")
        self.btn_stop.setObjectName("stopButton")
        self.btn_stop.setToolTip("请求停止当前任务。正在处理中的少量文件可能会先完成。")
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self.stop_process)
        btn_layout.addWidget(self.btn_start)
        btn_layout.addWidget(self.btn_stop)
        right_layout.addLayout(btn_layout, 0)

        log_group = QFrame()
        log_group.setObjectName("CardFrame")
        log_group.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self._add_shadow(log_group)
        log_layout = QVBoxLayout(log_group)
        log_layout.setSpacing(12)
        log_layout.setContentsMargins(16, 16, 16, 16)
        log_buttons = QHBoxLayout()
        log_buttons.setSpacing(10)
        log_buttons.setContentsMargins(0, 0, 0, 0)
        log_buttons.addWidget(self._section_title("", "处理日志"))
        log_buttons.addStretch()
        self.btn_clear_log = QPushButton("清空日志")
        self.btn_clear_log.setObjectName("logToolButton")
        self.btn_clear_log.setToolTip("清空当前界面日志，不删除已保存的日志文件。")
        self.btn_clear_log.clicked.connect(self._clear_log)
        self.btn_export_log = QPushButton("导出日志")
        self.btn_export_log.setObjectName("logToolButton")
        self.btn_export_log.setToolTip("把当前界面日志另存为文本文件。")
        self.btn_export_log.clicked.connect(self._export_log)
        log_buttons.addWidget(self.btn_clear_log)
        log_buttons.addWidget(self.btn_export_log)
        log_layout.addLayout(log_buttons)
        self.text_log = QTextEdit()
        self.text_log.setReadOnly(True)
        self.text_log.setMinimumHeight(180)
        self.text_log.setToolTip("运行日志。绿色表示完成，红色表示错误，橙色表示警告，紫色表示进度。")
        self.text_log.document().setMaximumBlockCount(LOG_MAX_BLOCK_COUNT)
        log_layout.addWidget(self.text_log, 1)
        right_layout.addWidget(log_group, 1)
        middle_layout.addLayout(right_layout, 7)
        main_layout.addLayout(middle_layout, 1)

        progress_group = QFrame()
        progress_group.setObjectName("CardFrame")
        progress_group.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self._add_shadow(progress_group, blur=12, y=2)
        progress_layout = QVBoxLayout(progress_group)
        progress_layout.setContentsMargins(14, 14, 14, 12)
        progress_layout.addWidget(self._section_title("", "总体进度"))
        self.progress_bar = AnimatedProgressBar()
        self.progress_bar.setToolTip("显示当前任务的大致进度。")
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        progress_layout.addWidget(self.progress_bar)
        progress_info_layout = QHBoxLayout()
        self.lbl_progress_detail = QLabel("准备就绪")
        self.lbl_progress_detail.setWordWrap(False)
        self.lbl_progress_detail.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        detail_h = self.lbl_progress_detail.sizeHint().height()
        self.lbl_progress_detail.setFixedHeight(detail_h)
        self.lbl_progress_detail.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.lbl_success_count = QLabel("成功: 0")
        self.lbl_failed_count = QLabel("失败: 0")
        self.lbl_skipped_count = QLabel("跳过: 0")
        for label in (self.lbl_progress_detail, self.lbl_success_count, self.lbl_failed_count, self.lbl_skipped_count):
            label.setStyleSheet(f"color: {self._colors['muted']};")
        progress_info_layout.addWidget(self.lbl_progress_detail, 3)
        progress_info_layout.addWidget(self.lbl_success_count)
        progress_info_layout.addWidget(self.lbl_failed_count)
        progress_info_layout.addWidget(self.lbl_skipped_count)
        progress_layout.addLayout(progress_info_layout)
        main_layout.addWidget(progress_group)

        self._apply_app_style()

        # 设置按钮图标（图标路径在 _apply_app_style 中创建）
        # 初始状态：开始按钮启用（白色图标），停止按钮禁用（灰色图标）
        if hasattr(self, '_play_ico_path') and self._play_ico_path:
            self.btn_start.setIcon(QIcon(self._play_ico_path))
            self.btn_start.setIconSize(QSize(20, 20))
        if hasattr(self, '_stop_ico_dis_path') and self._stop_ico_dis_path:
            self.btn_stop.setIcon(QIcon(self._stop_ico_dis_path))
            self.btn_stop.setIconSize(QSize(18, 18))
        elif hasattr(self, '_stop_ico_path') and self._stop_ico_path:
            self.btn_stop.setIcon(QIcon(self._stop_ico_path))
            self.btn_stop.setIconSize(QSize(18, 18))

    def _center_message_box(self, box):
        box.adjustSize()
        center = self.frameGeometry().center()
        box.move(center.x() - box.width() // 2, center.y() - box.height() // 2)

    def _question_box(self, title, text):
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Icon.Question)
        box.setWindowTitle(title)
        box.setText(text)
        box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        box.setDefaultButton(QMessageBox.StandardButton.No)
        self._center_message_box(box)
        return box.exec()

    def _info_box(self, title, text):
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Icon.Information)
        box.setWindowTitle(title)
        box.setText(text)
        box.setStandardButtons(QMessageBox.StandardButton.Ok)
        box.setDefaultButton(QMessageBox.StandardButton.Ok)
        self._center_message_box(box)
        return box.exec()

    def _clear_log(self):
        self.text_log.clear()
        self.progress_bar.setValue(0)
        self.lbl_progress_detail.setText("日志已清空")
        self.lbl_success_count.setText("成功: 0")
        self.lbl_failed_count.setText("失败: 0")
        self.lbl_skipped_count.setText("跳过: 0")

    def _export_log(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "导出日志", "处理日志.txt", "Text Files (*.txt);;All Files (*)"
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(self.text_log.toPlainText())
            self._info_box("导出完成", f"日志已导出到:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"无法导出日志:\n{e}")

    def _open_output_dir(self):
        output = self._last_output_dir or self.le_output.text().strip()
        if not output:
            return
        try:
            if sys.platform == 'win32':
                os.startfile(output)
            else:
                import subprocess
                subprocess.Popen(['open' if sys.platform == 'darwin' else 'xdg-open', output])
        except Exception as e:
            print(f"[警告] 打开输出目录失败: {e}")

    def _browse_dir(self, line_edit, is_source=False):
        path = QFileDialog.getExistingDirectory(self, "选择文件夹")
        if path:
            path = os.path.normpath(path)
            line_edit.setText(path)
            if is_source and not self.le_output.text().strip():
                self.le_output.setText(suggest_output_dir(path))

    def _browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel Files (*.xlsx *.xlsm)")
        if path:
            self.le_excel.setText(os.path.normpath(path))
            self._debounce_header_preview()

    def _debounce_header_preview(self):
        """防抖：用户停止输入 200ms 后再读取 Excel"""
        self._debounce_timer.start()

    def _do_update_header_preview(self):
        """启动后台线程读取 Excel 表头，并更新列下拉。"""
        excel_path = self.le_excel.text().strip()
        sheet_name = self.le_sheet.text().strip() or None
        col_text = self.cb_col.currentData() or "A"

        if not excel_path:
            self.lbl_header_preview.setText("表头：(未选择 Excel)")
            self.lbl_header_preview.setStyleSheet(f"color: {self._colors['muted']};")
            return

        if not col_text or not re.match(r'^[A-Z]+$', col_text):
            self.lbl_header_preview.setText("表头：(列输入有误)")
            self.lbl_header_preview.setStyleSheet(f"color: {self._colors['red']}; font-weight: bold;")
            return

        self.lbl_header_preview.setText("表头：(读取中...)")
        self.lbl_header_preview.setStyleSheet(f"color: {self._colors['muted']}; font-style: italic;")
        worker = ExcelHeaderWorker(excel_path, sheet_name, col_text)
        self.header_worker = worker
        self.header_workers.append(worker)
        worker.loaded.connect(self._on_excel_headers_loaded)
        worker.failed.connect(self._on_excel_headers_failed)
        worker.finished.connect(lambda w=worker: self._cleanup_excel_worker(w))
        worker.start()

    def _cleanup_excel_worker(self, worker):
        if worker in self.header_workers:
            self.header_workers.remove(worker)
        if self.header_worker is worker:
            self.header_worker = self.header_workers[-1] if self.header_workers else None

    def _is_current_excel_request(self, excel_path, sheet_name, col_text):
        return (
            excel_path == self.le_excel.text().strip()
            and sheet_name == (self.le_sheet.text().strip() or "")
            and col_text == (self.cb_col.currentData() or "A")
        )

    def _on_excel_headers_loaded(self, excel_path, sheet_name, col_text, headers, cell_val):
        if not self._is_current_excel_request(excel_path, sheet_name, col_text):
            return
        self._populate_col_combo(headers)
        current_col = self.cb_col.currentData() or "A"
        if current_col != col_text:
            self._debounce_header_preview()
            return
        if cell_val is not None:
            self.lbl_header_preview.setText(f"表头：{str(cell_val).strip()}")
            self.lbl_header_preview.setStyleSheet(f"color: {self._colors['green']}; font-weight: bold;")
        else:
            self.lbl_header_preview.setText("表头：(该列第一行为空)")
            self.lbl_header_preview.setStyleSheet(f"color: {self._colors.get('tip_text', '#F57C00')};")

    def _on_excel_headers_failed(self, excel_path, sheet_name, col_text, error):
        if not self._is_current_excel_request(excel_path, sheet_name, col_text):
            return
        self._populate_col_combo([])
        self.lbl_header_preview.setText(f"表头：(读取失败: {error})")
        self.lbl_header_preview.setStyleSheet(f"color: {self._colors['red']}; font-weight: bold;")

    def _advance_progress(self, text):
        """从日志文本中提取进度值并更新进度条"""
        m = re.search(r'进度.*?(\d+)%', text)
        if m:
            self.progress_bar.setValue(int(m.group(1)))
            return
        m = re.search(r'进度.*?(\d+)/(\d+)', text)
        if m:
            cur = int(m.group(1))
            total = int(m.group(2))
            if total > 0:
                self.progress_bar.setValue(int(cur / total * 100))
            return

    def _update_stats_from_log(self, text):
        """从日志消息中提取统计计数并更新界面标签。

        统计标签与日志格式的约定集中在此方法中，避免散落各处。
        支持显式 STAT 标签和遗留关键词两种格式。
        """
        if not hasattr(self, 'lbl_success_count'):
            return

        # 优先匹配显式 STAT 标签，格式固定不受日志措辞影响
        m = re.search(r'\[STAT success=(\d+) failed=(\d+) skipped=(\d+)\]', text)
        if m:
            self.lbl_success_count.setText(f"成功: {m.group(1)}")
            self.lbl_failed_count.setText(f"失败: {m.group(2)}")
            self.lbl_skipped_count.setText(f"跳过: {m.group(3)}")
            return

        # 遗留关键词匹配（保持与现有日志格式兼容）
        stat_patterns = [
            ('lbl_success_count', r'成功[:：]\s*(\d+)', '成功'),
            ('lbl_failed_count',  r'失败[:：]?\s*(\d+)', '失败'),
            ('lbl_skipped_count', r'跳过[:：]?\s*(\d+)', '跳过'),
        ]
        for attr, pattern, label in stat_patterns:
            m = re.search(pattern, text)
            if m:
                getattr(self, attr).setText(f"{label}: {m.group(1)}")

    def append_log(self, text):
        """追加彩色日志并更新进度条"""
        self._advance_progress(text)

        cursor = self.text_log.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)

        fmt = QTextCharFormat()
        is_dark = self._colors['window'] == '#111827'
        if "[错误]" in text or "[失败]" in text or "异常" in text:
            fmt.setForeground(QColor("#F87171" if is_dark else "#D32F2F"))
            fmt.setFontWeight(QFont.Weight.Bold)
        elif "[警告]" in text or "[提醒]" in text or "手动处理" in text:
            fmt.setForeground(QColor("#FBBF24" if is_dark else "#F57C00"))
        elif "[完成]" in text or "[结果]" in text or "全部处理完成" in text:
            fmt.setForeground(QColor("#4ADE80" if is_dark else "#388E3C"))
            fmt.setFontWeight(QFont.Weight.Bold)
        elif "[停止]" in text or "[预览]" in text:
            fmt.setForeground(QColor("#60A5FA" if is_dark else "#1976D2"))
        elif "进度" in text:
            fmt.setForeground(QColor("#C084FC" if is_dark else "#6A1B9A"))
        elif "[步骤" in text:
            fmt.setForeground(QColor("#93C5FD" if is_dark else "#5C6BC0"))
        else:
            fmt.setForeground(QColor(self._colors['text']))

        cursor.setCharFormat(fmt)
        cursor.insertText(text)

        self.text_log.setTextCursor(cursor)
        self.text_log.ensureCursorVisible()
        if hasattr(self, 'lbl_progress_detail') and text.strip():
            single_line = ' '.join(text.strip().split())
            self.lbl_progress_detail.setText(single_line)
        self._update_stats_from_log(text)

    def _update_mode_help(self, mode_name):
        """显示当前处理模式的说明，避免只靠下拉项猜含义。"""
        self.lbl_mode_help.setText(MODE_HELP_TEXT.get(mode_name, ""))

    def _set_processing_controls_enabled(self, enabled):
        widgets = [
            self.le_source,
            self.le_output,
            self.le_excel,
            self.cb_col,
            self.le_sheet,
            self.le_start,
            self.le_end,
            self.cb_mode,
            self.cb_resize,
            self.cb_force_format,
            self.cb_move,
            self.chk_open_output,
        ]
        for widget in widgets:
            widget.setEnabled(enabled)
        if enabled:
            self._update_excel_params_enabled()

    def start_process(self):
        source = self.le_source.text().strip()
        output = self.le_output.text().strip()

        if not source:
            QMessageBox.warning(self, "警告", "请选择源文件夹！")
            return
        if not output:
            QMessageBox.warning(self, "警告", "请选择输出目录！")
            return

        source_norm = os.path.normpath(source)
        output_norm = os.path.normpath(output)
        if source_norm == output_norm:
            output = suggest_output_dir(source_norm)
            self.le_output.setText(output)
            output_norm = os.path.normpath(output)
            print(f"[提示] 输出目录不能与源目录相同，已自动改为: {output}\n")
        try:
            validate_output_dir(source_norm, output_norm)
        except ValueError as e:
            QMessageBox.warning(self, "警告", str(e))
            return
        self._last_output_dir = output

        col_text = self.cb_col.currentData() or "A"

        mode_map = {"完整流程": 1, "预览模式": 2, "仅分类(无Excel)": 4}
        mode_val = mode_map.get(self.cb_mode.currentText(), 1)
        resize_map = {"crop - 铺满裁剪": "crop", "stretch - 拉伸": "stretch", "fit - 留白缩放": "fit"}
        resize_val = resize_map.get(self.cb_resize.currentText(), "crop")

        start_row = self.le_start.text().strip()
        start_row = int(start_row) if start_row.isdigit() else None

        end_row = self.le_end.text().strip()
        end_row = int(end_row) if end_row.isdigit() else None

        copy_mode = (self.cb_move.currentIndex() == 1)  # True=复制, False=剪切

        force_format_map = {"保持原格式": None, "JPG": "jpg", "PNG": "png"}
        force_format_val = force_format_map.get(self.cb_force_format.currentText(), None)

        kwargs = {
            'source_dir': source,
            'output_root': output,
            'mode': mode_val,
            'excel_path': self.le_excel.text().strip() or None,
            'col': col_text or 'A',
            'sheet_name': self.le_sheet.text().strip() or None,
            'start_row': start_row,
            'end_row': end_row,
            'force_format': force_format_val,
            'resize_mode': resize_val,
            'stop_event': self.stop_event,
            'copy_mode': copy_mode,
        }

        self.text_log.clear()
        self.progress_bar.setFormat("%p%")
        self.progress_bar.setValue(0)
        self.lbl_progress_detail.setText("正在准备处理...")
        self.lbl_success_count.setText("成功: 0")
        self.lbl_failed_count.setText("失败: 0")
        self.lbl_skipped_count.setText("跳过: 0")
        self.stop_event.clear()

        self._set_processing_controls_enabled(False)
        self.btn_start.setEnabled(False)
        self.btn_start.setIcon(QIcon(self._play_ico_dis_path) if hasattr(self, '_play_ico_dis_path') and self._play_ico_dis_path else QIcon())
        self.btn_start.setIconSize(QSize(20, 20))
        self.btn_stop.setEnabled(True)
        self.btn_stop.setIcon(QIcon(self._stop_ico_path) if hasattr(self, '_stop_ico_path') and self._stop_ico_path else QIcon())
        self.btn_stop.setIconSize(QSize(18, 18))

        self.worker = WorkerThread(kwargs)
        self.worker.finished_signal.connect(self._process_finished)
        self.worker.start()
        self.progress_bar.start_animation()

    def stop_process(self):
        if self.worker and self.worker.isRunning():
            reply = self._question_box("确认停止", "当前任务正在处理图片。\n\n确定要停止当前任务吗？")
            if reply != QMessageBox.StandardButton.Yes:
                return
            print("\n>>> 正在发送停止信号，请稍等...")
            self.stop_event.set()

    def _process_finished(self):
        self.progress_bar.stop_animation()
        self._set_processing_controls_enabled(True)
        self.btn_start.setEnabled(True)
        self.btn_start.setIcon(QIcon(self._play_ico_path) if hasattr(self, '_play_ico_path') and self._play_ico_path else QIcon())
        self.btn_start.setIconSize(QSize(20, 20))
        self.btn_stop.setEnabled(False)
        self.btn_stop.setIcon(QIcon(self._stop_ico_dis_path) if hasattr(self, '_stop_ico_dis_path') and self._stop_ico_dis_path else QIcon())
        self.btn_stop.setIconSize(QSize(18, 18))

        completed = not self.stop_event.is_set()
        if completed:
            self.progress_bar.setValue(100)
            self.lbl_progress_detail.setText("处理完成")
        else:
            self.progress_bar.setFormat("已停止 %p%")
            self.lbl_progress_detail.setText("任务已停止")

        print("\n[系统] 后台任务已结束。")
        QApplication.alert(self, 0)
        if completed:
            try:
                if sys.platform == 'win32':
                    winsound.MessageBeep(winsound.MB_ICONASTERISK)
                else:
                    QApplication.beep()
            except Exception:
                QApplication.beep()
            if hasattr(self, 'chk_open_output') and self.chk_open_output.isChecked():
                self._open_output_dir()
            self._info_box("处理完成", "全部处理完成！\n\n请在输出目录查看处理结果。")

    def _restore_stdout(self):
        if getattr(self, '_stdout_stream', None) is not None and sys.stdout is self._stdout_stream:
            sys.stdout = self._old_stdout
        self._stdout_stream = None

    def closeEvent(self, event):
        # 清理临时图标文件
        for p in _TEMP_ICON_PATHS:
            try:
                os.remove(p)
            except OSError:
                pass
        _TEMP_ICON_PATHS.clear()
        for header_worker in list(getattr(self, 'header_workers', [])):
            if header_worker.isRunning():
                header_worker.requestInterruption()
                header_worker.wait(1000)
        if self.worker and self.worker.isRunning():
            reply = QMessageBox.question(
                self, '确认退出',
                '当前有图片正在处理中，强制退出可能会导致部分文件处理中断。\n\n确定要中止并退出吗？',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                print("\n[系统] 正在中止任务并准备退出...")
                self.stop_event.set()
                self.worker.wait(5000)
                if self.worker.isRunning():
                    QMessageBox.warning(self, "仍在处理中", "后台任务仍未停止，请稍后再退出。")
                    event.ignore()
                    return
                self._restore_stdout()
                event.accept()
            else:
                event.ignore()
        else:
            self._restore_stdout()
            event.accept()


# ============================================================
# 入口
# ============================================================

# 嵌入的 ICO 图标（base64 编码）
APP_ICON_B64 = "AAABAAUAEBAAAAAAIAA8AgAAVgAAACAgAAAAACAAKQQAAJICAAAwMAAAAAAgAGoGAAC7BgAAQEAAAAAAIACDCAAAJQ0AAAAAAAAAACAArAYAAKgVAACJUE5HDQoaCgAAAA1JSERSAAAAEAAAABAIBgAAAB/z/2EAAAIDSURBVHicdZO9T1RREMV/5967yy4fLqCNqAQhGrUwWmmBsdHCxFBa2JiovR2JlY0Jkc7E0r/ARhsi2tOgjcS4GldCwMqIycoG9rHv3bF4y2aBdZK5uZ9nzj0zIzCBDODMnE1aoLS7S08rFgEjqc3qR75jEphOPmWsGHhpMG0ZRQz1RBAmR0ue5TTj4fpjrYon5iaKvCtUuJHWAfGf12DtwQ9BusXy0WNMa3zOJl2kKuEByVArgtmB4ILgOssIeHNcCibKGEJ4DEsjGu2HUsgjCsMrsp16/myDz+l5hJnRH5TlzJzgbxNunoUHV2B1Ewoe0ig2dzyXj8OrT/DmM1TKEA0RsdBNsRVhfBjeVuHFknFqRJwY2ODa2CKvf99nfNiTxv0aha45AnYz6AsQzZi/LbYaDZr1GjuJ0UzzL3VDOA5Y8PCrAZWyY+YCXJ06z8LPZxzpC0QDHcjRIQZJCrfOwfWpXJepUeP5jLFWdyxUD6e2AxAjDJXg/TdoJOAcLH7NYb0TBQ8f1mGwlN/dIxLMI9I8ZU6w3cqVhhwEy88MGChCULugwHAo4Gi2lYmAnGCkfJgqQLR9eA7RdBsTrAHLfhC3B51Zb++U8iDOIiuFfr447iiLjnvZDkt40o6avRzAk2UJH7PI3dojJepu59PzdtFE2ZLe/aQ+rGAk32dZAUXM9A/HsdHyrVMXaAAAAABJRU5ErkJggolQTkcNChoKAAAADUlIRFIAAAAgAAAAIAgGAAAAc3p69AAAA/BJREFUeJy1l09oXFUUxn/n3vdmJp2kYNMYkZaJJrGkKkor1aILEVFBs6iYujIoCFpQoavatBJDNLEEUVduRARdaOpGcFH8uxCUKFUXtlhCbVKSohhoJP9n3rvHxZ3pJJMw8xKmH8xi3rvvnO+c891z7oVK9KhFVdY9rwdG1VY+WuuoRy2nJQbIDWubOHJqEBxbI2RQAGuY/uuYjAPQr4YBcesJjKrlsMRtg/qISXNSY+41AaktOa6AKxBj+V3zjEy8Jp+tDlRWs2p7Q1+xGd4DcCv+23oQAIyEICFEi5yaPCGvlkhIKfJbB/UhGvhW8zgUB1gqS7R1KIpDUJsliBbpneiTj3tG1RrOoaCiwpCUFkJQR+cAgmBRxOVxOIZy/Zo5fRhnGBCXGyaHYX8x7euUWkcaVgtgUuzSgH0gagBw3GJCAkCpb+QbwUmAGqEDwCT5whQpaYK1pTWmehjX3ga1DAqwWICUBU3CABCB5QgyNa3XIGAE5vPwRBeceBia0mVS650qquAQlgvw/o/w4c/QmAZXhXhVAiKwUoCn74btaTj1PWTCciaMeONGlKVIyASwkIcXDkLvPfDBWDXrCQhAuaZjl2HkO2hugrjYnpYKsC2E5UjouOEK/y618vdVS3uz8mC7JNJMIhEKkA6guRF2Zv3vxkY4ctARBnD0wJecfbGbzw8dZ0dTRCbw2UuCBDLxUIXIgTpYzEPbDhjpVi5ehV3ZC4Tbr3BXyx8YKRBrgCTc0YkJlGCMT333XgiMYeRxeOqjI8zGzYxNHSDSBqxo4omeqASrETu/G568E0DY0wIvPbCNvjPPsrt5L492wnxeMAktb6oEgYGFAtzfBre1+B0gAs/fp8wtOR7rFPq+Mon7RWICJUHNrfimdOgOr4fSbhj/R9h/s2V2Cab/89szaT+vScAIzCz4XnD2KFiBrtbix8U072mFrpvAWvi0Fzp3wk+TnkQtIlUJOIVsGt79wUe2PeMNfj1ebkYiYA2gvmekLJz5E744B6mg9ommKgFVSBmYmoXBb8otdXWKK8utRV1kQ0iHtedHok6YCqAlrLVyLZxWnwGJCUCxCW1C2ZuBAdCtHru3DkW9TwNgLVOugMOX9jrFeg2GGFHDZf+nX82l37iojvMSQvFQer3gJIB4hZlsnl88gdsRfz7ndQkRBEf97gNrnCtEJoMRZfj8gMz3jKpdczHJvalvhVmOuRXQyAu5TuIwYjGmAaI5Ppk4Kc+ULyYllC4oQ/qcWvpE6JBNz8qNoTGoYxrHO5eOy9s+YBSkYmYWM9HxsqZ1N/tcTLvGm5+Ya2ABw2R6hl8vjMgcqoJIFaFvcIWuGzaw/T9hSn05HagaDwAAAABJRU5ErkJggolQTkcNChoKAAAADUlIRFIAAAAwAAAAMAgGAAAAVwL5hwAABjFJREFUeJzNmm2IXNUZx3/POXdedrOpujHZWGNejUpedFMIpGqbaPtJMVQlUeyLgoiISEQ/qBiYnaJiKUXaQoS2IFVB3fRD8IMipC+hitJGjIIBk0hMosS8aMxmd3Z37r3n8cO5d2ZNZrPOnZ3d/cOyM/eec8///5znPM8z51zh+6CkhpXI+o+R79W+Rezy/xxlFESzP6lfLapTQnpclDSA8Tk0vqEq9CGUxQEsek4X2yorEOa5CFXXnpkQgxpBEAac8PFnj8kntZslNSmf8wsY03Dp03qXWu5DWScBRbHtoN0ADlxIhPCROF4qfsC2vdulyia1bJd4fAEJ+QUl7c518rzJsVlj0BBQnIJrtz8pIIogWMmByUM8wrsMc8/Bsuw7eybqfEpqABbBD0yRf9sivdEQkYAgmHPEth8KqCqxLZJzMSfiChsOl2XvWBHmO13K4iTP32wHvdEQVRECBDsN5EnGNCLk4hFCY5lr8/Rf/XudBZAGFy+gpAFlcUue0k22k9ujIUIR8tNAuiFqIrpYOVDlCcri6MNCatnEfRbn2GUKXOdGceItP5PgxILGHLddXHlgiwyAikn9aVnApSKs1SoyA8kDGI3AFJgfDbAOgBLW8B/vRs5yjclTQInP+5jphOCwqAhrAdYDZv0Gfy+O6UmWawupu71IiIkq89NrtShkLNHUU8oGMXWuQfqh2fLASGIRJXuQTfoK4JqZd62PGJyv3XgQgaGqF1ET0gLCGGZlDNpNCzAJ+RuXw8M/gbyFwEzcrxFi5y3/yh54cbcX0dRM0KSAdKo7cvDszXDZBc0NNh56L4X/HoTDp6AYNCei6RlIBRQtDIdw96uw/6QfWMcZWAGDYpIaTDGETujKwwt3wKKLoDPXvPUzCUhFWAOVKrx3yLtCo7Ug1Nd46IRK6PNjR+CNsK8CR07D4m6Izqn02yigRlC8345EYM8SkLqbEaUaCwtmn+LxH28jZ0N+9+4DfD7Yw+yCkjNSaz/lAkhIqia1L34mzozCT5fCL1bBlh1KV0H508/6uHb5W6CGJbM/57YdzzMYtTw8GePHeR4oUI1gwzL45Y+Ua5cYYhfR2/MJUWUOUeViVsw5wIWFUSKXOlkL400O7TrCGObNhltW+JL9uY2O0OX5w//vIcgPExQG2LbnNxw500XeuLE5KRNan8MxsAKnRuDW5bDwIqjGsGyO4bEb4PE37uTD471AzL8OrWRZN3xxunX7TaoAxC/s2672X41ArPDgdbBzn7Lzs6twCptWKb2XCKV/tv5Tb9JcSMRHoyXdfgGDnxHBZ+tnbhIK1rFpleOR64UTQzAaeZGtoOUZSOuhwMA3w3DLSh9aY+dzRcrvmh/CnzcauoswEvprqcBW0JIAVRgc9UloNPKEb1/t70nCTAHnYN+X0DPLtwtjOFmBwWo9gWWNRZkEWOMt3FWAny+HA1/5fLB6Plwx15MSvP8r8OlxOD4IgQVroZiHnIUbLveLHSBnsonIVI0Oh956c7vghTsn7nPlfP83Fn9cWP/s1Ce/IIM/NSUgzbSVEB59HR663lvSJqFAte46YzFeuoqdv/7aHjh0agrKaahXo+8dhndeTgingzZjwbTKS/pmIQ8Z14CqL38nE1nIQwtRKOuAk41Jr4WmGjUBYmbuftA5GLN3VROgbkZuJzaG1rnWN7aEARy0XN+2HwqcTr+YXeAAJOBDVyVmZm7sAsnJjUOc8H7tGuB/edxPsHghu02e1RriYMYJUQRUOSNwxcEn5Riq4l2oD8tfJBT4qxQQVTLuEbQPqkSmEwH+cfBJOUa/WkTS81cVSsgKCCp53rZF1sbDhCJMcrrKCCWWAKvKiSBmzf6QowCUxSWL2J+G7y1LNRLuciEnbZGcKiE6reFVVYkkwGJQjfjV/q3yBXvrZ9gNj1mX/lZXU+TvJs8aNwIa4VSm1q1EEQzWdoCrcjQe4d5DJXmTfrVsrp8Vj3vQfUlJOzs62aqOX5scC6b0oDXZZHIhXwvskDP0ffq0HDmbfGMB8N3T+mf1AgfrDKxRoUejxDrtgKBJZjptLB+EVf53eKt4f29AfgKo0K/TH0oneOFkYkuqCtsxU/WqTYr6KzfnvuAxFt8C9D1gI7weTUkAAAAASUVORK5CYIKJUE5HDQoaCgAAAA1JSERSAAAAQAAAAEAIBgAAAKppcd4AAAhKSURBVHic7Zt/iB1XFcc/5868H7ub3aRr0pSsabKJaZNu1KXVpOTXUlIjtLQFcVsrQiGGIgXF1KrUWF42qSlqxFhBqP9o0z/EDaKgQQoGG4q0gi02P0piJWtqY7Mbkibdn/PezD3+ceft2337dptk5+2v9guzO29m7p17vvecc+89cy58yCHXXCKnhhak7eR1lK0ijragnETpQEE04epV6FQP1Rkl9LjIqX+1bf3gh3Jq6BBb/Llyvy61IasJWYeQEkVVpk8bRFF8wHDMS3HqrSb+xYMSAdCp3vD5eOUnrL1YwaOaWtnMw2r4ilo2GJ86fFc4aV27bkRgC0QCJ8XjBe3n4JkO6fkgEsYnIKc+HRLe/LTe7qf5haRYTwi2AChWwY5bdhogiiB44oNJgw34n43Y+Z9d0lmuxaPKVawtFr75ad0uGQ6IUB8NEQoIghm33PRDAVXFGh9ffLABB7p2yU6nCdhyBzlWkFhllu/VR7xafq0BqCVC8KZKioRgUdSrx4t6ea7r+/K1SuYwmoBYVZp/oJ8Sn1dUyRJC3OuzD4oihKaWlPax/cxT8qtyEkqCxcPGspxmgefFo1ZDdNYKD85gFV+HiDTFz1fs0VW0Y8npsEzDJ2278egQ66X5gqmlNQoIZfap/VgIYiPUZKjD8F1ElJaS5pdMIGalOcURk6HNBthZaPeVoah4oMpFCWg50yE9bvYisSrEtr/EpwnDnVoAmMWqXw5BNMKaDAs1y2YADjn5nJCxSqSF2yRFVp3zmKlD3XVBBcWgYrkDoLiWMSN/iOEz4sNMm+QkAVGECFEcAUdjGcvVPD3lLZt6jJJxNAE6c6b2VUOZjHPH0V0n/CQrM1L9FaIAUYIvSIwAIzBQgNC6c5TklkzFutSdzsuAJkRCIgQYgb483NEE39icbAPL8dt/QucbUJ8Bm8A7Jk2ACOQjaJoPL3wZGmsn36iJsKkZLg3AX96ChszkzWHSBBhgsABrFzvhCxaiCF4/53posj5BBMIIbrsJFtSAb2BjMxw+5e5N1uEkYgKCE1yBlIGOF+HAyzA/e409pGDEIqJYNagKxsBAHjYsh9894uoPwuTcS2JOUCg16vQFmJd2x3gESPzwSF/hiWUgNBQiyPqQ8SwWQ8aHs5cdETWpZMNRVZkHZHyn/uMdiuvFfOTO3TXlvcBwyw3/ZevKV1hYc4UrgQGUyELaK5GWJBKdBxQx0Qgg4oRf0Qj9eejphxrPcjkwbP/kH9lz1z7S/iDdvUvYcfgnvHZ+DWnPolqdOduUzwSNQF8AT26F3DboHVICa1i76G2euWsfaaAw2MjihnM8u3UvNak8URW/x0wpASJQiGBJA6xbCve3wJYVyoV+WLngXSQ1SKFQg+8VCIP5NNWdpzE7SGhdbKsamFICPIHeALbdCo11zv5/eK/QWKO83r2KS+83kaq9hCj4dT38/Xwr3f3zyHgWrVJ4YkoJsOoc5Jda46EzgjWLhSfalOPdjez4835O9azhsk3x0ul7+PqRHKH1sFU0gao4wUow4oax1Yvhs0udo/SNI+Wr6w1/OKH89e013N/5G+oyl7k0sIDePLTcCBf6JZFpb8V2VafaCi8SCCJ4oAU84+YHRpwZ1Gdg3z1CjW+JgMuDC8hHyuZlyuObIOUCmlUxgikjILRQn4UH1sYvjqXxxN3bsBwevdPw3gAMRcrdnxB2bxV6A7g4AL5XnWX2lBDgxUPfxmVw8wKn9kZG37cK32qDlQth6wrhsfVOY94PHEHV8gJV8QFjGisuAvlQqxN8OGZQvB0vauoz8LP7YDCAoRAynjuqGZ6uigZEsYPzjbPfQuRmfnevcve9ChKJwOnzEORLztEz0F9w/6E6MYbENGBk22rT0NPnBAdnww+1uuuRLQlUxGAeTpyDC73O1sGRaBVOdsO7V6CpAdL+2HdNFokQUFwGg2t47nOw9ibXq4Kz5Ydb4x6UUk8WI119ATTUwKL60j2LU/8vfhqWfQy23QJ16dJcIilMuirFrdTeueJ+ewI33wA7t1R+fuTHxqIlLKp3RyV8vBE+v3r0tTMX3XuS0IRJE2DVrdFPnIdv/wke3+J6qDwmamS04xsJl9Yx+nnFaVBxCV2s4/fHXVywIevMabJIRJmsukDowX/A4TdLBCQNVejudYQnheScoLog5UDBRYirAaEUcU6K4ETnAZE62yz38kki6TVB4hMhHf4zO/Ch/zb4EQHT3YDpxigCpFqBt5mEMhlHEaA6R7LCJkKZjAbizQYOx9StvedUghSMSJKCEwBtjMwSa49Hr4jjmqfAXPQNGn+9M7w28rITVMSiKrURXUSclhSgcypTTI1gbJ6BPLwMFbLE2nbjvdkheTU8L2lkTqXKKVayCAWOvvM9+Tft6hX3DwwTcHQ3EapiAg7aAS4YH4+5QoJB1YL1+RGo0D7yVhEiyiGMy6PlMfFdPtJsT51TpeDV4OsQB84+KS/RjhmZLj/uhonmvfpTr4FvRn2EKN5sTJ1VpeDXkYoGeLVrFZs4BHTiMjBijPX2D2LJqd/1lOyM+nnO1OBjEFXCWaINihIB1q8jFQX8rZDnXtqx5cLDuOO9CjmEDrHNz+gO8dkvPvN1CDQiUkGlWl8rrxMqbn+ICJ7JglrQkANdXXyHX0qhmB5fXm5iIeI0+uV79FaT5gngPkmxGANMuBtvGhDrsh0iEMOLRDx7ZpccARhPeLiaGd+IPTYrfqw3omxUy+2ErANSxMwnJMa1Q9zbxecN4FhU4NWzu+QUAO3qcWjsTrFrR04NnTp71gk5NbRfXXuvredUhUOYtpPI0Zk3Ryg6dDveJsmPUAH/B52PTyocgfy7AAAAAElFTkSuQmCCiVBORw0KGgoAAAANSUhEUgAAAQAAAAEACAYAAABccqhmAAAGc0lEQVR4nO3dzZEbVRiFYQ1FIvaKINiQAilQRRD22gRBFSmQAhuCYGWHMiwo1ZjxzEjdfX++2+d5tlC4Lem8fSVrzOUCAAAAAAAAAAAAAAAAANT2MPsCenv36fFx9jWwti8fH067k1P9xoydUc4ShaV/EwZPFasGYcmLNnyqWi0Ey1ys0bOaFWJQ/gINn9VVDkHZCzN8zqZiCL6bfQEvMX7OqOLrulSRKj5A0EOV00CZE4Dxk6TK671EAKo8GDBShdf91GNIhQcAKpj1lmDaCcD44cmsPUwJgPHDt2bsYngAjB9eN3ofQwNg/HDbyJ0MC4Dxw/1G7WVIAIwfthuxmxLfAwDm6B4Ad3/Yr/d+ugbA+OG4njvqFgDjh3Z67clnABCsSwDc/aG9HrtqHgDjh35a78tbAAjWNADu/tBfy505AUCwZgFw94dxWu3NCQCCCQAEaxIAx38Yr8XunAAg2OEAuPvDPEf35wQAwQQAggkABDsUAO//Yb4jO3QCgGACAMEEAIIJAATbHQAfAEIde/foBADBBACCCQAEEwAIJgAQTAAgmABAMAGAYAIAwQQAggkABBMACCYAEEwAIJgAQDABgGACAMEEAIIJAAQTAAgmABBMACCYAEAwAYBg38++gDP4/GH2FeR6/9vsK1ibABxg+PNdnwMh2MdbgJ2MvxbPxz4CsIMXW02el+0EYCMvsto8P9sIAAQTgA3cXdbgebqfAEAwAYBgAgDBfBGoM19QOc57+n4EoBPDb+f6WApBe94CQDAB6MDdvw+Pa3sCAMEEAIIJAAQTAAgmABBMACCYAEAwAYBgvgpME59//fHVf/b+978HXglbCACHvDX85/+OENTjLQC73TP+I/8+/QlAkJY/Tbd3zCJQiwCEaRGBoyMWgToEIESru3+r8YpADQIQyF+swZUABHhp8Hsi0Pqu7RQwnwBAMAEI5q0AAnByRs5bBCCcQGQTAEQgmACcmGFziwBwuVzui0XrH+bxw0HzCcBJ7fpzfieGOALAJq3u2u7+NQgA/zPirYDx1yEAJ3T0KN8zAsZfiwCw29YxG389/kqwk2n2Y78f7vufcV5H7e8EXJMA0ISRr8lbAF7ljwXPTwBOpMdgR0bgj5/H/Vr8RwAowfjnEABu6n0KMP55BOAkeo+013/f+OcSAKYx/vn8MeAJjPqg7t7vBtxi+HU4AbDJ0dgYfy0CwDC3xv/Ln2OugycCsLgZX9bZ82u689ckAOyyJQLGX5cA0JXx1yYAC5v9Xf1bv77x1ycAdGH8axCARc2++1+9dB3Gvw5fBOKw6xeEDH89TgA0YfxrEoAFVTn+X/31z+wrYC8B4BDjX5sALKbS3d/41ycA7GL85yAAbGb85yEAC6lw/O81fj8JOIcAcDd3/vPxRSBuMvzzcgJYxKzjv/GfmwB0UOG9egvVxn+Wx7USAVjAjBd+tfHTh88AOrmOtsXfojtatfG78/cjAJ2t+OL96Yfxv+aKj9MZeAsAwQQAggkABBOADVb8QC+R5+l+AgDBBGAjd5faPD/bCMAOXmQ1eV62E4CdvNhq8Xzs44tAB1xfdL7EMo/hHyMADXgRsipvASCYAEAwAYBgAgDBBACCCQAEEwAIJgAQTAAgmABAMAGAYAIAwQQAggkABBMACCYAEEwAIJgAQDABgGACAMEEAIIJAAQTAAgmABBMACCYAEAwAYBgAgDBBACCCQAEEwAIJgAQTAAgmABAMAGAYAIAwXYH4MvHh4eWFwLst3ePTgAQTAAgmABAMAGAYIcC4INAmO/IDp0AIJgAQDABgGCHA+BzAJjn6P6cACBYkwA4BcB4LXbnBADBBACCNQuAtwEwTqu9OQFAsKYBcAqA/lruzAkAgjUPgFMA9NN6X11OACIA7fXYlbcAEKxbAJwCoJ1ee+p6AhABOK7njrq/BRAB2K/3fnwGAMGGBMApALYbsZthJwARgPuN2svQtwAiALeN3MnwzwBEAF43eh9TPgQUAfjWjF1M+1MAEYAns/ZQYoTvPj0+zr4GmGH2jbDE9wBmPwgwQ4XXfYkAXC41HgwYpcrrvcRFPOctAWdVZfhXZU4AX6v2IEELFV/X5S7oOacBVldx+FdlL+w5IWA1lYd/Vf4CXyIGVLXC6L+21MU+JwRUsdrwr5a86NcIAqOsOvjnTvGbeIsocNRZxg4AAAAAAAAAAAAAAAAs7F/PVZZepLdhWgAAAABJRU5ErkJggg=="

def _remove_temp_icon(icon_path):
    try:
        if os.path.exists(icon_path):
            os.remove(icon_path)
    except OSError:
        pass


def main():
    if sys.platform == 'win32':
        import ctypes
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(2)
        except Exception:
            try:
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    # Qt6 默认启用高 DPI 缩放，不再需要 QT_ENABLE_HIGHDPI_SCALING 环境变量
    font = QFont(["Microsoft YaHei", "Segoe UI", "Noto Sans CJK SC"])
    font.setPointSize(10)
    font.setStyleStrategy(
        QFont.StyleStrategy.PreferAntialias | QFont.StyleStrategy.PreferQuality
    )
    # 高 DPI 下 PreferNoHinting 渲染最清晰，PreferDefaultHinting 会使用位图 hint 导致锯齿
    font.setHintingPreference(QFont.HintingPreference.PreferNoHinting)
    app.setFont(font)

    # 从嵌入的多尺寸 ICO 设置窗口图标，避免任务栏只拿到低分辨率图层。
    icon_bytes = base64.b64decode(APP_ICON_B64)
    icon_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".ico") as f:
            icon_path = f.name
            f.write(icon_bytes)
        app.setWindowIcon(QIcon(icon_path))
        app.aboutToQuit.connect(lambda p=icon_path: _remove_temp_icon(p))
    except Exception:
        if icon_path:
            _remove_temp_icon(icon_path)
        pm = QPixmap()
        pm.loadFromData(icon_bytes)
        app.setWindowIcon(QIcon(pm))

    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()

