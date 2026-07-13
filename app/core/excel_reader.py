# -*- coding: utf-8 -*-
"""Excel 读取 & 条码解析"""

import os
import re

from app.constants import EXCEL_EMPTY_ROW_BREAK_THRESHOLD
from app.core.file_ops import unique_preserve_order


def excel_col_to_index(col):
    col = (col or "").strip().upper()
    if not re.match(r'^[A-Z]+$', col):
        raise ValueError(f"非法列名: {col}")
    col_idx = 0
    for ch in col:
        col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
    return col_idx


def step3_read_excel(excel_path, col, log, start_row=None, end_row=None, sheet_name=None, stop_event=None):
    if not os.path.isfile(excel_path):
        log.write(f"[错误] Excel文件不存在: {excel_path}")
        return []
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                log.write(f"[错误] 工作表 '{sheet_name}' 不存在，可用: {wb.sheetnames}")
                return []
            ws = wb[sheet_name]
        else:
            ws = wb.active
        col_idx = excel_col_to_index(col)
        auto_end_row = end_row is None
        start_row = start_row or 2
        if end_row is not None and end_row < start_row:
            log.write(f"[错误] 行范围无效: {start_row} ~ {end_row}")
            return []
        if end_row is None:
            log.write(f"  行范围: {start_row} ~ 自动")
        else:
            log.write(f"  行范围: {start_row} ~ {end_row}")
        barcodes = []
        if end_row is not None:
            total = end_row - start_row + 1
        else:
            max_row = ws.max_row or 0
            total = max_row - start_row + 1 if max_row >= start_row else None
        last_pct = -1
        empty_streak = 0
        stopped_early = False
        for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row,
                                             min_col=col_idx, max_col=col_idx, values_only=True)):
            if stop_event and stop_event.is_set():
                log.write("  [停止] 用户中止Excel读取")
                break
            val = row[0] if row else None
            if val is not None:
                if isinstance(val, float) and val.is_integer():
                    val = int(val)
                bc = str(val).strip()
                if bc:
                    barcodes.append(bc)
                    empty_streak = 0
                else:
                    empty_streak += 1
            else:
                empty_streak += 1
            if auto_end_row and empty_streak >= EXCEL_EMPTY_ROW_BREAK_THRESHOLD:
                stopped_early = True
                log.write(
                    f"  [信息] 连续 {EXCEL_EMPTY_ROW_BREAK_THRESHOLD} 行为空，"
                    "已提前结束读取"
                )
                break
            if total:
                pct = int((i + 1) / total * 100)
                if pct >= last_pct + 20:
                    last_pct = pct
                    log.write(f"  读取进度: {pct}% ({i + 1}/{total})")
            elif (i + 1) % 1000 == 0:
                log.write(f"  读取进度: 已扫描 {i + 1} 行")
        barcodes, dup = unique_preserve_order(barcodes)
        if stopped_early:
            log.write("  [信息] 如需读取空行后的数据，请手动填写结束行")
        log.write(f"  [完成] 读取 {len(barcodes)} 个条码（去重 {dup} 个）")
        return barcodes
    finally:
        wb.close()


def read_excel_preview(excel_path, sheet_name=None, col_text='A'):
    """读取 Excel 表头列表和当前列表头。供 UI 后台线程调用。

    sheet_name 指定但不存在时抛出 ValueError，与 step3_read_excel 行为一致。
    """
    if not excel_path or not os.path.isfile(excel_path):
        return [], None
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在，可用: {wb.sheetnames}")
            ws = wb[sheet_name]
        else:
            ws = wb.active
        headers = []
        max_col = ws.max_column or 0
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val is not None:
                label = str(val).strip()
                if label:
                    col_letter = ""
                    n = col_idx
                    while n > 0:
                        n, rem = divmod(n - 1, 26)
                        col_letter = chr(65 + rem) + col_letter
                    headers.append((col_letter, label))
        cell_val = None
        if col_text and re.match(r'^[A-Z]+$', col_text):
            cell_val = ws.cell(row=1, column=excel_col_to_index(col_text)).value
        return headers, cell_val
    finally:
        wb.close()